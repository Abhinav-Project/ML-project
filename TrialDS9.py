import streamlit.components.v1 as components
import streamlit as st
import pandas as pd
from pulp import LpProblem, LpMinimize, LpVariable, lpSum, value
import plotly.express as px
import time
import numpy as np
from scipy.interpolate import interp1d
import bisect
import random
from scipy.spatial import ConvexHull
from io import BytesIO
from openpyxl import Workbook

# Define the optimization function with usability constraints
def optimize_mixture_core(w1, w2, materials, cost, yield_factors, scope1, scope3, composition, 
                          chemistry_constraints, inv_min, inv_max, material_group, group_constraints, 
                          scrap_flags, usability_constraints, custom_groups, LM_target,
                          emission_constraint=None):
    prob = LpProblem("Steel_Optimization", LpMinimize)
    x = {mat: LpVariable(f"x_{mat}", lowBound=0) for mat in materials}

    if not materials:
        return {}, 0, 0

    # Safely calculate max_cost_val
    relevant_costs = [c for m, c in cost.items() if m in materials and pd.notna(c)]
    max_cost_val = max(relevant_costs) if relevant_costs else 1.0
    if max_cost_val == 0: max_cost_val = 1.0


    # Safely calculate max_emission_val
    all_emissions = []
    for mat in materials:
        if mat in scope1 and mat in scope3 and pd.notna(scope1[mat]) and pd.notna(scope3[mat]):
            all_emissions.append(scope1[mat] + scope3[mat])
    max_emission_val = max(all_emissions) if all_emissions else 1.0
    if max_emission_val == 0: max_emission_val = 1.0
    
    # Objective Function
    if emission_constraint is None:
        obj_terms = []
        for mat in materials:
            cost_val = cost.get(mat, 0)
            s1_val = scope1.get(mat, 0)
            s3_val = scope3.get(mat, 0)
            term = x[mat] * (w1 * (cost_val / max_cost_val) + w2 * ((s1_val + s3_val) / max_emission_val))
            obj_terms.append(term)
        if obj_terms: prob += lpSum(obj_terms)
    else:
        cost_terms = [x[mat] * cost.get(mat, 0) for mat in materials]
        if cost_terms: prob += lpSum(cost_terms)

    # Yield Constraint (normalized to produce 1 ton of LM)
    yield_terms = [x[mat] * yield_factors.get(mat, 0) for mat in materials]
    if yield_terms: prob += lpSum(yield_terms) == 1

    # Chemistry Constraints
    for element, bounds in chemistry_constraints.items():
        min_bound = bounds.get("min", 0)
        max_bound = bounds.get("max", float('inf'))
        recovery = bounds.get("recovery", 0) / 100
        
        chem_expr_terms = []
        for mat in materials:
            comp_val = composition.get(mat, {}).get(element, 0)
            yield_val = yield_factors.get(mat, 0)
            chem_expr_terms.append(x[mat] * comp_val * yield_val * recovery)
        
        if chem_expr_terms:
            chem_expr = lpSum(chem_expr_terms)
            if pd.notna(min_bound) and min_bound > -float('inf'):
                 prob += chem_expr >= min_bound, f"{element}_min"
            if pd.notna(max_bound) and max_bound < float('inf'):
                 prob += chem_expr <= max_bound, f"{element}_max"

    # Inventory Constraints (normalized by LM_target)
    for mat in materials:
        if inv_min.get(mat, 0) > 0:
            prob += x[mat] >= inv_min[mat] / LM_target, f"{mat}_inv_min"
        if inv_max.get(mat, 0) > 0 and inv_max.get(mat,0) < float('inf'):
            prob += x[mat] <= inv_max[mat] / LM_target, f"{mat}_inv_max"

    # Predefined Group Constraints (normalized by LM_target)
    for group, bounds in group_constraints.items():
        group_mats = [mat for mat in materials if material_group.get(mat) == group]
        if group_mats:
            group_sum = lpSum(x[mat] for mat in group_mats)
            group_min = bounds.get("Group Min", 0)
            group_max = bounds.get("Group Max", 0)
            
            if group_min > 0:
                prob += group_sum >= group_min / LM_target, f"{group}_group_min"
            if group_max > 0 and group_max < float('inf'):
                prob += group_sum <= group_max / LM_target, f"{group}_group_max"
    
    # Custom Group Constraints (normalized by LM_target)
    for group_name, group_data in custom_groups.items():
        group_mats = [mat for mat in group_data["materials"] if mat in materials]
        if group_mats:
            group_sum = lpSum(x[mat] for mat in group_mats)
            group_min = group_data.get("min", 0)
            group_max = group_data.get("max", 0)
            
            if group_min > 0:
                prob += group_sum >= group_min / LM_target, f"custom_{group_name}_min"
            if group_max > 0 and group_max < float('inf'):
                prob += group_sum <= group_max / LM_target, f"custom_{group_name}_max"
    
    # Scrap Constraint
    scrap_materials = [mat for mat in materials if scrap_flags.get(mat, 0) == 1]
    if materials: 
        total_sum_materials = lpSum(x[mat] for mat in materials)
        if scrap_materials:
            prob += lpSum(x[mat] for mat in scrap_materials) >= 0.75 * total_sum_materials
        elif 0.75 > 0:
            prob += 0 >= 0.75 * total_sum_materials
    
    # Usability Constraints (normalized by LM_target)
    for mat in materials:
        constraints = usability_constraints.get(mat, {"min": 0, "max": 0})
        min_val = constraints.get("min", 0)
        max_val = constraints.get("max", 0)
        
        if min_val > 0:
            prob += x[mat] >= min_val / LM_target, f"{mat}_usability_min"
        if max_val > 0 and max_val < float('inf'):
            prob += x[mat] <= max_val / LM_target, f"{mat}_usability_max"
    
    # Emission constraint for ε-constraint method
    if emission_constraint is not None:
        emission_terms = [x[mat] * (scope1.get(mat,0) + scope3.get(mat,0)) for mat in materials]
        if emission_terms:
            prob += lpSum(emission_terms) <= emission_constraint, "emission_constraint"

    prob.solve()

    mix_out = {mat: value(x[mat]) * 1000 for mat in materials if value(x[mat]) > 1e-6}
    cost_out = sum((value(x[mat])) * cost.get(mat,0) for mat in materials if value(x[mat]) > 1e-6)
    emission_out = sum((value(x[mat])) * (scope1.get(mat,0) + scope3.get(mat,0)) for mat in materials if value(x[mat]) > 1e-6)
    
    return mix_out, cost_out, emission_out

def select_gap(solutions, max_cost=None):
    if len(solutions) < 2:
        return None
    
    solutions = sorted(solutions, key=lambda x: x[2])
    gaps = []
    slopes = []
    
    min_emission = min(sol[2] for sol in solutions)
    max_emission = max(sol[2] for sol in solutions)
    
    if max_cost is None:
        max_cost = max(sol[1] for sol in solutions)
    min_cost = min(sol[1] for sol in solutions)
    
    for i in range(1, len(solutions)):
        emission_diff = (solutions[i][2] - solutions[i-1][2]) / (max_emission - min_emission + 1e-6)
        cost_diff = (solutions[i][1] - solutions[i-1][1]) / (max_cost - min_cost + 1e-6)
        gap_size = np.sqrt(emission_diff**2 + cost_diff**2)
        
        if solutions[i][2] - solutions[i-1][2] != 0:
            slope = (solutions[i][1] - solutions[i-1][1]) / (solutions[i][2] - solutions[i-1][2])
        else:
            slope = float('inf')
        
        gaps.append(gap_size)
        slopes.append(slope)
    
    slope_changes = []
    for i in range(1, len(slopes)):
        if slopes[i] != float('inf') and slopes[i-1] != float('inf'):
            slope_change = abs(slopes[i] - slopes[i-1])
        else:
            slope_change = 0
        slope_changes.append(slope_change)
    
    if len(slope_changes) < len(gaps):
        slope_changes = [0] + slope_changes
    
    weights = []
    for i in range(len(gaps)):
        gap_weight = 0.7 * gaps[i] + 0.3 * (slope_changes[i] if i < len(slope_changes) else 0)
        weights.append(gap_weight)
    
    total_weight = sum(weights)
    if total_weight > 0:
        normalized_weights = [w / total_weight for w in weights]
    else:
        normalized_weights = [1.0 / len(weights)] * len(weights)
    
    if len(normalized_weights) > 0:
        selected_gap_idx = random.choices(range(len(normalized_weights)), weights=normalized_weights, k=1)[0]
        selected_epsilon = (solutions[selected_gap_idx][2] + solutions[selected_gap_idx+1][2]) / 2
        return selected_epsilon
    else:
        return None

def explore_nonconvex(solutions, n_explore=5):
    if len(solutions) < 3:
        return []
    
    solutions = sorted(solutions, key=lambda x: x[2])
    nonconvex_points = []
    slopes = []
    
    for i in range(1, len(solutions)):
        if solutions[i][2] - solutions[i-1][2] != 0:
            slope = (solutions[i][1] - solutions[i-1][1]) / (solutions[i][2] - solutions[i-1][2])
            slopes.append(slope)
    
    for i in range(1, len(slopes)):
        if slopes[i] * slopes[i-1] < 0:
            lower_idx = i-1
            upper_idx = i+1
            emission_range = solutions[upper_idx][2] - solutions[lower_idx][2]
            
            for j in range(1, n_explore+1):
                position_ratio = j/(n_explore+1)
                random_factor = 0.1 * emission_range * (random.random() - 0.5)
                new_epsilon = (solutions[i][2] + 
                              position_ratio * (solutions[i+1][2] - solutions[i][2]) + 
                              random_factor)
                new_epsilon = max(solutions[i][2], min(solutions[i+1][2], new_epsilon))
                nonconvex_points.append(new_epsilon)
    
    if not nonconvex_points and len(solutions) >= 2:
        for _ in range(n_explore):
            idx = random.randint(0, len(solutions)-2)
            base = (solutions[idx][2] + solutions[idx+1][2]) / 2
            random_factor = 0.1 * (solutions[-1][2] - solutions[0][2]) * (random.random() - 0.5)
            nonconvex_points.append(base + random_factor)
    
    return nonconvex_points[:n_explore*2]  

def should_terminate(solutions, iteration, max_iterations=100, max_solutions=100,min_improvement=0.005, min_gap_threshold=0.0005):
    if iteration >= max_iterations:
        return True
    if len(solutions) >= max_solutions:
        return True
    if len(solutions) >= 2:
        solutions_sorted = sorted(solutions, key=lambda x: x[2])
        min_gap = min(solutions_sorted[i+1][2] - solutions_sorted[i][2] 
                    for i in range(len(solutions_sorted)-1))
        if min_gap < min_gap_threshold:
            return True
    if len(solutions) >= 3:
        solutions_sorted = sorted(solutions, key=lambda x: x[2])
        cost_range = max(sol[1] for sol in solutions_sorted) - min(sol[1] for sol in solutions_sorted)
        emission_range = max(sol[2] for sol in solutions_sorted) - min(sol[2] for sol in solutions_sorted)
        
        area = 0
        for i in range(1, len(solutions_sorted)):
            area += (solutions_sorted[i][2] - solutions_sorted[i-1][2]) * (solutions_sorted[i-1][1] - solutions_sorted[i][1])
        
        if hasattr(should_terminate, 'prev_area'):
            improvement = (area - should_terminate.prev_area) / should_terminate.prev_area
            if improvement < min_improvement:
                return True
        
        should_terminate.prev_area = area
    
    return False

def adaptive_e_constraint_method(min_emission, max_emission, max_iterations=30, max_solutions=30):
    solutions = []
    epsilon_values = []
    
    initial_epsilons = list(np.linspace(min_emission, max_emission, 10))
    intermediate_epsilons = list(np.linspace(min_emission, max_emission, 10))[1:-1:2]
    edge_epsilons = [0.8 * min_emission + 0.2 * max_emission,
                     0.2 * min_emission + 0.8 * max_emission]
    
    all_initial_epsilons = list(set(initial_epsilons + intermediate_epsilons + edge_epsilons))
    all_initial_epsilons.sort()
    
    for epsilon in all_initial_epsilons:
        mix, cost_val, emission_val = optimize_mixture_core(
            1, 0, current_materials, current_cost, current_yield_factors,
            current_scope1, current_scope3, current_composition, current_chemistry_constraints,
            current_inv_min, current_inv_max, current_material_group, current_group_constraints,
            current_scrap_flags, current_usability_constraints, current_custom_groups, current_LM_target,
            emission_constraint=epsilon
        )
        if mix:
            solutions.append((mix, cost_val, emission_val))
            epsilon_values.append(epsilon)
    
    max_cost = max(sol[1] for sol in solutions) if solutions else 0
    
    for iteration in range(max_iterations):
        if should_terminate(solutions, iteration, max_iterations):
            break
        
        selected_epsilon = select_gap(solutions, max_cost)
        exploration_epsilons = explore_nonconvex(solutions)
        
        epsilons_to_try = [selected_epsilon] if selected_epsilon is not None else []
        epsilons_to_try.extend(exploration_epsilons)
        
        remaining_slots = max_solutions - len(solutions)
        epsilons_to_try = epsilons_to_try[:remaining_slots]
        
        for epsilon in epsilons_to_try:
            mix, cost_val, emission_val = optimize_mixture_core(
                1, 0, current_materials, current_cost, current_yield_factors,
                current_scope1, current_scope3, current_composition, current_chemistry_constraints,
                current_inv_min, current_inv_max, current_material_group, current_group_constraints,
                current_scrap_flags, current_usability_constraints, current_custom_groups, current_LM_target,
                emission_constraint=epsilon
            )
            
            if mix:
                insert_pos = bisect.bisect_left(epsilon_values, epsilon)
                solutions.insert(insert_pos, (mix, cost_val, emission_val))
                epsilon_values.insert(insert_pos, epsilon)
    
    return solutions

def create_excel_download(solutions):
    """Create an Excel file with all solutions in separate sheets"""
    wb = Workbook()
    
    # Create summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Headers for solution summary
    summary_headers = [
        "Solution", "Method", "Cost (₹/ton)", "Emissions (tCO₂)", 
        "Scrap %", "Pre-Consumer %", "Post-Consumer %", 
        "Recycle Content %", "Yield %", "Total Input (kg)"
    ]
    
    # Write headers to summary sheet
    ws_summary.append(summary_headers)
    
    # Prepare data for the Excel sheet
    for i, sol in enumerate(solutions):
        # Add solution summary
        solution_row = [
            f"Solution {i+1}",
            sol["Method"],
            sol["Cost (INR/ton)"],
            sol["Emissions (tCO₂/tcs)"],
            sol["Scrap %"],
            sol["Pre-Consumer %"],
            sol["Post-Consumer %"],
            sol["Recycle Content %"],
            sol["Yield %"],
            sol["Total Input (kg)"]
        ]
        ws_summary.append(solution_row)
    
    # Create charge mix sheet in matrix format
    ws_mix = wb.create_sheet("Charge Mixes kg per ton")  # Changed sheet title to remove invalid character
    
    # Get all materials from all solutions
    all_materials = set()
    for sol in solutions:
        all_materials.update(sol["mix"].keys())
    all_materials = sorted(all_materials)
    
    # Write headers - first column is material, then solution numbers
    header_row = ["Material"] + [f"Solution {i+1}" for i in range(len(solutions))]
    ws_mix.append(header_row)
    
    # Write each material's quantities across solutions
    for mat in all_materials:
        row = [mat]
        for sol in solutions:
            row.append(sol["mix"].get(mat, 0))
        ws_mix.append(row)
    
    # Create a sheet for solution details
    ws_details = wb.create_sheet("Solution Details")
    
    # Write detailed information for each solution
    detail_headers = [
        "Solution", "Method", "Cost (₹/ton)", "Emissions (tCO₂)", 
        "Scrap %", "Pre-Consumer %", "Post-Consumer %", 
        "Recycle Content %", "Yield %", "Total Input (kg)", "Total Input (ton)"
    ]
    ws_details.append(detail_headers)
    
    for i, sol in enumerate(solutions):
        detail_row = [
            f"Solution {i+1}",
            sol["Method"],
            sol["Cost (INR/ton)"],
            sol["Emissions (tCO₂/tcs)"],
            sol["Scrap %"],
            sol["Pre-Consumer %"],
            sol["Post-Consumer %"],
            sol["Recycle Content %"],
            sol["Yield %"],
            sol["Total Input (kg)"],
            sol["Total Input (ton)"]
        ]
        ws_details.append(detail_row)
    
    # Format columns
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
    
    # Save to BytesIO object
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

def create_heat_template(solutions, composition_df):
    """Create a heat template Excel file in the specified format"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Heat Template"
    
    # Get all material columns from the original template
    material_columns = [
        "201 Scrap Import", "304 Processed (DOM)", "304 Processed (IMP)", 
        "316 Filter Chips (DOM)", "Al -Bronze Scrap - Revert", "Alloyed Briquettes/Turnings", 
        "Aluminium Bar", "Ca Mo", "Charge Chrome Fines (50%)", "Charge Cr (L.P) Imp (51%)", 
        "Circle Cutting", "Cu Ni 70/30 Imported", "Cu Pure", "Cu Pure Scrap", "CuNi Scrap IMP", 
        "Fe Boron", "Fe Moly", "Fe Ni (MC) DOM 21.5%", "Fe Ni High Carbon (Ni 23.30%)", 
        "Fe Ni LC 20%", "Fe Ni LC 22%", "Fe Vanadium", "Fe-Nb", "FERRO NICKEL NI 15-18%", 
        "FeSi 70%", "FeSi 75% (Higrade)", "FeSi 75% (Imp)", "FeSi fines (0-3 mm)", 
        "FeTi (30-32%)", "Grinding Dust (300 series with Mo)", "Grinding Dust (300 series)", 
        "HC Fe Cr Low Co Chrome", "HC Fe Cr Sized", "HC FeCr - chips (3-10 mm)", 
        "HC FeCr - fine (0-10 mm)", "HC FeCr - fine (0-3 mm)", "HC FeCr HP 60%", 
        "HC FeCr Low Silicon", "HC FeCr LP 60%", "HC FeMn HP (P=0.3%)", "HC FeMn MP (P=0.15%)", 
        "High Ni Scrap", "LC SiMn", "LCFeCr 65%", "LMS Bundles", "Maruti Bundle", 
        "MC SiMn", "Meatball Cu 5-6%", "Meatball Cu-15%", "Meatball Cu-20%", 
        "Mix of HMS and LMS", "Mn metal", "Mo Oxide", "MS SCRAP HP END CUT", 
        "MS Scrap HP HMS", "MS Scrap HP HMS Imp", "MS SCRAP LP BUNDLEIMPORT", 
        "MS SCRAP LP BUSHLING DOM", "MS SCRAP LP BUSHLING IMP", "MS SCRAP NUT BOLT, PLATE, BABRI", 
        "MS Shredded Scrap", "MS Shredded Scrap (Dom)", "MS Turning (Dom)", "MS Turning (Imp)", 
        "MS Wire Bundle", "MS Wire Loose", "Ni Pig Iron 11%", "NICKEL PIG IRON  NI-14%", 
        "NICKEL PIG IRON  NI-17%", "Ni-Pure", "SAF Metal", "Sponge Iron (Coal based)", 
        "SS 301 Scrap Domestic", "SS 304 Domestic (Higher TCO)", "SS 304 Domestic Turning", 
        "SS 316 Domestic (Higher TCO)", "SS 316 Domestic Turning", "SS 400 Series Purchase HP", 
        "SS 400 Series Purchase LP", "SS Filter Ni8Cr13Cu12", "SS Filter Ni8Cr13Cu12 Imported", 
        "SS High Ni High Cu Imported", "SS Ingot (10/16/2)", "SS Ingot (18/8)", 
        "SS MIX SOLID SCRAP 201", "SS Scrap (300 series with Mo)", "SS Scrap (300 series)", 
        "SS Scrap (400 series)", "SS Scrap (Blade Steel)", "SS Scrap (Blooms)", 
        "SS Scrap (J3/Jsl)", "SS Scrap 18/10", "SS Scrap 18/8 Local", 
        "SS Scrap 304 Imported (Near Shore)", "SS Scrap 304 Sized Dom", "SS Scrap 310 S 13/25", 
        "SS SCRAP 310, IMP", "SS Scrap 316 Imported (Far Shore)", "SS Scrap 316 Imported (Near Shore)", 
        "SS Scrap 316 Local", "SS Scrap 316 Tuning (Imp)", "SS Scrap 317 Type (12.25/25)", 
        "SS Scrap Cu Ni Ingot DOM", "SS Scrap Cu Ni TB DOM", "SS Scrap Duplex Grade", 
        "SS Scrap High Ni", "SS Scrap J1 Purchase", "SS Scrap J4 Purchased Cr>!2", 
        "SS Scrap Low Ni/Sorted (Dom) J4 Cr>12", "SS Scrap Ni4Cr12Cu3", "SS Scrap Ni5Cr13Cu5", 
        "SS Scrap Piston Rings", "SS Utencils (14.0/2.9)", "SS Utensil Induction Grade", 
        "SS201 SCRAP,VIETNAM", "Ti cored Wire (98%)"
    ]
    
    # Write headers
    headers = [
        "Heat_id", "grade", "furnace_type", "C_frac", "Mn_frac", "S_frac", "P_frac", 
        "Si_frac", "Ni_frac", "Cr_frac", "Cu_frac", "Mo_frac", "Al_frac", "V_frac", 
        "B_frac", "Nb_frac", "Ti_frac", "W_frac", "total_charge_wt_kg", "lime_total_kg", 
        "dolomite_total_kg", "coke_total_kg", "oxygen_per_t_Nm3", "vcb_on_time_min", 
        "tap_temperature_C"
    ] + material_columns + ["target_power_kwh_per_t"]
    
    ws.append(headers)
    
    # Add each solution as a row
    for i, sol in enumerate(solutions):
        mix = sol["mix"]
        total_input = sum(mix.values())
        
        # Initialize element sums (in kg)
        element_sums = {
            "C": 0, "Mn": 0, "S": 0, "P": 0, "Si": 0, "Ni": 0, "Cr": 0,
            "Cu": 0, "Mo": 0, "Al": 0, "V": 0, "B": 0, "Nb": 0, "Ti": 0, "W": 0
        }
        
        # Calculate total element contributions from each material
        for mat, qty in mix.items():
            # Get composition for this material from the composition_df
            # composition_df is a DataFrame with materials as index and elements as columns
            if mat in composition_df.index:
                mat_composition = composition_df.loc[mat]
                
                # Add each element's contribution (convert from % to kg)
                for element in element_sums.keys():
                    if element in mat_composition:
                        element_content = mat_composition[element]
                        if pd.notna(element_content):
                            element_sums[element] += qty * (element_content / 100)  # Convert % to kg
        
        # Calculate element fractions (as percentages of total input)
        element_fractions = {}
        for element, total_kg in element_sums.items():
            if total_input > 0:
                element_fractions[element] = (total_kg / total_input) * 100  # Convert back to %
            else:
                element_fractions[element] = 0
        
        # Create row data
        row_data = [
            f"Solution {i+1}",  # Heat_id
            st.session_state.selected_grade,  # grade
            st.session_state.selected_route,  # furnace_type
            element_fractions.get("C", 0),  # C_frac
            element_fractions.get("Mn", 0),  # Mn_frac
            element_fractions.get("S", 0),  # S_frac
            element_fractions.get("P", 0),  # P_frac
            element_fractions.get("Si", 0),  # Si_frac
            element_fractions.get("Ni", 0),  # Ni_frac
            element_fractions.get("Cr", 0),  # Cr_frac
            element_fractions.get("Cu", 0),  # Cu_frac
            element_fractions.get("Mo", 0),  # Mo_frac
            element_fractions.get("Al", 0),  # Al_frac
            element_fractions.get("V", 0),  # V_frac
            element_fractions.get("B", 0),  # B_frac
            element_fractions.get("Nb", 0),  # Nb_frac
            element_fractions.get("Ti", 0),  # Ti_frac
            element_fractions.get("W", 0),  # W_frac
            total_input,  # total_charge_wt_kg
            0,  # lime_total_kg
            0,  # dolomite_total_kg
            0,  # coke_total_kg
            0,  # oxygen_per_t_Nm3
            0,  # vcb_on_time_min_per_t
            0   # tap_temperature_C
        ]
        
        # Add material quantities
        for mat_col in material_columns:
            row_data.append(mix.get(mat_col, 0))
        
        # Add target power (empty)
        row_data.append(0)  # target_power_kwh_per_t
        
        ws.append(row_data)
    
    # Format columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to BytesIO object
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

def create_recycled_content_pie_chart(pre_consumer_pct, post_consumer_pct):
    """Create a pie chart for recycled content breakdown"""
    labels = ['Pre-Consumer', 'Post-Consumer', 'Ferro Alloys & Virgin Materials']
    values = [pre_consumer_pct, post_consumer_pct, 100 - (pre_consumer_pct + post_consumer_pct)]
    
    fig = px.pie(
        values=values,
        names=labels,
        color=labels,
        color_discrete_map={
            'Pre-Consumer': '#1f77b4',
            'Post-Consumer': '#ff7f0e',
            'Ferro Alloys & Virgin Materials': '#2ca02c'
        },
        hole=0.3,
        title="Recycled Content Breakdown"
    )
    
    fig.update_traces(
        textposition='inside',
        textinfo='percent+label',
        hovertemplate="%{label}: %{value:.1f}%<extra></extra>"
    )
    
    fig.update_layout(
        showlegend=True,
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=-0.2,
            xanchor="center",
            x=0.5
        ),
        margin=dict(t=50, b=20, l=20, r=20),
        height=350
    )
    
    return fig

def create_component_emissions_pie_chart(mix, scope1, scope3):
    """Create a pie chart showing component-wise emissions breakdown"""
    component_emissions = {}
    
    for material, quantity_kg in mix.items():
        # Convert kg to tons for emission calculation
        quantity_tons = quantity_kg / 1000
        material_emission = (scope1.get(material, 0) + scope3.get(material, 0)) * quantity_tons
        if material_emission > 0:
            component_emissions[material] = material_emission
    
    # Sort by emissions and take top components
    sorted_components = sorted(component_emissions.items(), key=lambda x: x[1], reverse=True)
    
    # Take top 10 components and group the rest as "Others"
    top_components = sorted_components[:10]
    other_emissions = sum(emission for _, emission in sorted_components[10:])
    
    labels = [f"{mat}" for mat, emission in top_components]
    values = [emission for mat, emission in top_components]
    
    if other_emissions > 0:
        labels.append("Other Materials")
        values.append(other_emissions)
    
    # Create a DataFrame for Plotly Express
    df = pd.DataFrame({
        'Material': labels,
        'Emissions (tCO₂)': values
    })
    
    # Create the pie chart
    fig = px.pie(
        df,
        values='Emissions (tCO₂)',
        names='Material',
        title="Component-wise Emissions Breakdown",
        hover_data=['Emissions (tCO₂)']
    )
    
    fig.update_traces(
        textposition='inside',
        textinfo='percent+label',
        hovertemplate="<b>%{label}</b><br>Emissions: %{value:.4f} tCO₂<br>Percentage: %{percent}",
        marker=dict(line=dict(color='#FFFFFF', width=1))
    )
    
    fig.update_layout(
        showlegend=True,
        legend=dict(
            orientation="v",
            yanchor="top",
            y=0.99,
            xanchor="left",
            x=1.05
        ),
        margin=dict(t=50, b=20, l=20, r=150),
        height=400,
        width=600
    )
    
    return fig

# Configure page settings
st.set_page_config(
    page_title="Stainless Steel Charge Mix Optimization",
    page_icon="⚙️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for corporate styling
st.markdown("""
    <style>
        :root {
            --primary-color: #005f87;
            --secondary-color: #003d5b;
            --accent-color: #00a0e1;
            --light-bg: #f5f9fa;
            --dark-text: #2c3e50;
            --light-text: #7f8c8d;
            --success-color: #28a745;
            --warning-color: #ffc107;
            --danger-color: #dc3545;
        }
        
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            color: var(--dark-text);
            background-color: white;
        }
        
        .main {
            max-width: 1200px;
            padding: 1rem;
        }
        
        .stButton>button {
            background-color: var(--primary-color);
            color: white;
            padding: 0.5rem 1rem;
            border-radius: 4px;
            border: none;
            font-weight: 500;
            transition: all 0.3s;
        }
        
        .stButton>button:hover {
            background-color: var(--secondary-color);
            transform: translateY(-1px);
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        
        .stButton>button:active {
            transform: translateY(0);
        }
        
        .info-box {
            background-color: var(--light-bg);
            padding: 1rem;
            border-radius: 6px;
            border-left: 4px solid var(--accent-color);
            margin: 0.75rem 0;
            font-size: 0.95rem;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        }
        
        .success-box {
            background-color: #e8f5e9;
            padding: 1rem;
            border-radius: 6px;
            border-left: 4px solid var(--success-color);
            margin: 0.75rem 0;
            font-size: 0.95rem;
        }
        
        .warning-box {
            background-color: #fff8e1;
            padding: 1rem;
            border-radius: 6px;
            border-left: 4px solid var(--warning-color);
            margin: 0.75rem 0;
            font-size: 0.95rem;
        }
        
        .error-box {
            background-color: #ffebee;
            padding: 1rem;
            border-radius: 6px;
            border-left: 4px solid var(--danger-color);
            margin: 0.75rem 0;
            font-size: 0.95rem;
        }
        
        .metric-card {
            background-color: var(--light-bg);
            padding: 1rem;
            border-radius: 6px;
            margin: 0.5rem 0;
            font-size: 0.95rem;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
            border: 1px solid #e0e0e0;
        }
        
        .compact-table {
            width: 100%;
            border-collapse: collapse;
            font-size: 0.9rem;
        }
        
        .compact-table th, .compact-table td {
            padding: 8px;
            text-align: center;
            border: 1px solid #e0e0e0;
        }
        
        .compact-table th {
            background-color: var(--primary-color);
            color: white;
            font-weight: 500;
        }
        
        .compact-table tr:nth-child(even) {
            background-color: var(--light-bg);
        }
        
        .element-control {
            margin-bottom: 0.75rem;
        }
        
        .element-name {
            font-weight: 600;
            color: var(--dark-text);
            margin-bottom: 0.3rem;
        }
        
        .default-value {
            font-size: 0.8rem;
            color: var(--light-text);
            margin-top: 0.2rem;
        }
        
        .changed-value {
            color: var(--danger-color);
            font-weight: 600;
        }
        
        .edit-button {
            background-color: var(--accent-color) !important;
        }
        
        .lock-button {
            background-color: var(--success-color) !important;
        }
        
        .reset-button {
            background-color: var(--danger-color) !important;
        }
        
        .running-animation {
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            padding: 2rem;
            background-color: rgba(245, 249, 250, 0.9);
            border-radius: 8px;
            margin: 1.5rem 0;
            box-shadow: 0 4px 8px rgba(0,0,0,0.1);
            border: 1px solid var(--accent-color);
        }
        
        .spinner {
            width: 50px;
            height: 50px;
            border: 5px solid #f3f3f3;
            border-top: 5px solid var(--accent-color);
            border-radius: 50%;
            animation: spin 1s linear infinite;
            margin-bottom: 1.5rem;
        }
        
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
        
        .progress-text {
            margin-top: 1.5rem;
            font-size: 1rem;
            color: var(--dark-text);
            text-align: center;
        }
        
        .progress-bar {
            width: 100%;
            height: 10px;
            background-color: #f3f3f3;
            border-radius: 5px;
            margin-top: 1rem;
            overflow: hidden;
        }
        
        .progress-bar-fill {
            position: absolute;
            top: 0;
            left: 0;
            height: 100%;
            background-color: var(--accent-color);
            width: 0%;
            transition: width 0.3s ease;
        }
        
        .section-header {
            color: var(--primary-color);
            border-bottom: 2px solid var(--accent-color);
            padding-bottom: 0.75rem;
            margin-top: 2rem;
            font-size: 1.5rem;
            font-weight: 600;
        }
        
        .chem-input {
            font-size: 0.9rem;
            padding: 0.4rem;
        }
        
        .header-container {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 1.5rem;
        }
        
        .button-group {
            display: flex;
            gap: 0.75rem;
        }
        
        .chem-row {
            display: flex;
            gap: 1.25rem;
            margin-bottom: 0.75rem;
            align-items: center;
            padding: 0.75rem;
            background-color: var(--light-bg);
            border-radius: 6px;
        }
        
        .chem-col {
            flex: 1;
        }
        
        .flux-fuel-table {
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 1.5rem;
        }
        
        .flux-fuel-table th, .flux-fuel-table td {
            padding: 10px;
            border: 1px solid #e0e0e0;
            text-align: left;
        }
        
        .flux-fuel-table th {
            background-color: var(--primary-color);
            color: white;
        }
        
        .flux-fuel-table tr:nth-child(even) {
            background-color: var(--light-bg);
        }
        
        .emission-breakdown {
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 1.5rem;
        }
        
        .emission-breakdown th, .emission-breakdown td {
            padding: 10px;
            border: 1px solid #e0e0e0;
            text-align: center;
        }
        
        .emission-breakdown th {
            background-color: var(--primary-color);
            color: white;
        }
        
        .emission-breakdown tr:nth-child(even) {
            background-color: var(--light-bg);
        }
        
        .highlight-row {
            font-weight: 600;
            background-color: #e6f3ff !important;
        }
        
        .sub-section {
            background-color: var(--light-bg);
            padding: 1.25rem;
            border-radius: 8px;
            margin-bottom: 1.5rem;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        }
        
        .sub-section-header {
            color: var(--primary-color);
            font-weight: 600;
            margin-bottom: 1rem;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        
        .route-tabs {
            margin-bottom: 1.5rem;
        }
        
        .summary-table {
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 1.5rem;
        }
        
        .summary-table th, .summary-table td {
            padding: 12px;
            border: 1px solid #e0e0e0;
            text-align: center;
        }
        
        .summary-table th {
            background-color: var(--primary-color);
            color: white;
        }
        
        .summary-table tr:nth-child(even) {
            background-color: var(--light-bg);
        }
        
        .material-constraint-row {
            display: flex;
            gap: 1.25rem;
            margin-bottom: 0.75rem;
            align-items: center;
            padding: 0.75rem;
            background-color: var(--light-bg);
            border-radius: 6px;
        }
        
        .material-constraint-col {
            flex: 1;
        }
        
        .constraint-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 1rem;
        }
        
        .usability-constraints-container {
            background-color: var(--light-bg);
            padding: 1.25rem;
            border-radius: 8px;
            margin-bottom: 1.5rem;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        }
        
        .group-container {
            background-color: var(--light-bg);
            padding: 1.25rem;
            border-radius: 8px;
            margin-bottom: 1.5rem;
            border-left: 4px solid var(--accent-color);
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        }
        
        .group-header {
            font-weight: 600;
            color: var(--primary-color);
            margin-bottom: 1rem;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        
        .custom-group-container {
            background-color: #e6f3ff;
            padding: 1.25rem;
            border-radius: 8px;
            margin-bottom: 1.5rem;
            border-left: 4px solid var(--accent-color);
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        }
        
        .custom-group-header {
            font-weight: 600;
            color: var(--primary-color);
            margin-bottom: 1rem;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        
        .group-constraint-row {
            display: flex;
            gap: 1.25rem;
            margin-bottom: 0.75rem;
            align-items: center;
        }
        
        .group-constraint-col {
            flex: 1;
        }
        
        .material-selector {
            max-height: 300px;
            overflow-y: auto;
            border: 1px solid #e0e0e0;
            padding: 0.75rem;
            border-radius: 6px;
            margin-bottom: 1rem;
            background-color: white;
        }
        
        .material-selector-item {
            padding: 0.5rem 0;
            display: flex;
            align-items: center;
        }
        
        .material-selector-checkbox {
            margin-right: 0.75rem;
        }
        
        .constraint-summary {
            font-size: 0.9rem;
            color: var(--light-text);
            margin-top: 0.5rem;
        }
        
        .constraint-error {
            color: var(--danger-color);
            font-size: 0.85rem;
            margin-top: 0.5rem;
        }
        
        .unconstrained-label {
            color: var(--light-text);
            font-style: italic;
        }
        
        .back-button {
            background-color: var(--warning-color) !important;
            margin-right: 0.75rem;
        }
        
        .workflow-container {
            width: 100%;
            margin-bottom: 2rem;
        }
        
        .workflow-steps {
            display: flex;
            justify-content: space-between;
            position: relative;
            margin-bottom: 1rem;
        }
        
        .workflow-step {
            display: flex;
            flex-direction: column;
            align-items: center;
            position: relative;
            z-index: 2;
            flex: 1;
            padding: 0 15px;
        }
        
        .step-icon {
            width: 36px;
            height: 36px;
            border-radius: 50%;
            display: flex;
            align-items: center;
            justify-content: center;
            margin-bottom: 0.75rem;
            background-color: #e0e0e0;
            color: var(--light-text);
            font-size: 1.1rem;
            transition: all 0.3s;
        }
        
        .step-label {
            font-size: 0.85rem;
            text-align: center;
            color: var(--light-text);
            white-space: nowrap;
            transition: all 0.3s;
        }
        
        .step-active .step-icon {
            background-color: var(--accent-color);
            color: white;
            transform: scale(1.1);
        }
        
        .step-completed .step-icon {
            background-color: var(--success-color);
            color: white;
        }
        
        .step-active .step-label,
        .step-completed .step-label {
            color: var(--primary-color);
            font-weight: 500;
        }
        
        .progress-line {
            position: absolute;
            top: 18px;
            left: 0;
            right: 0;
            height: 4px;
            background-color: #e0e0e0;
            z-index: 1;
        }
        
        .progress-fill {
            position: absolute;
            top: 0;
            left: 0;
            height: 100%;
            background-color: var(--success-color);
            transition: width 0.3s ease;
        }
        
        .loading-container {
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            padding: 3rem;
            background: linear-gradient(135deg, #f5f9fa 0%, #e4e8eb 100%);
            border-radius: 12px;
            box-shadow: 0 8px 16px rgba(0,0,0,0.1);
            margin: 2.5rem 0;
            border: 1px solid var(--accent-color);
        }
        
        .loading-title {
            font-size: 1.75rem;
            color: var(--primary-color);
            margin-bottom: 1.5rem;
            text-align: center;
            font-weight: 600;
        }
        
        .loading-subtitle {
            font-size: 1.1rem;
            color: var(--light-text);
            margin-bottom: 2rem;
            text-align: center;
        }
        
        .loading-spinner {
            width: 70px;
            height: 70px;
            border: 8px solid rgba(0, 160, 225, 0.2);
            border-top: 8px solid var(--accent-color);
            border-radius: 50%;
            animation: spin 1.5s linear infinite;
            margin-bottom: 2rem;
        }
        
        .loading-progress-container {
            width: 100%;
            max-width: 500px;
            margin-top: 1.5rem;
        }
        
        .loading-progress-text {
            font-size: 1rem;
            color: var(--primary-color);
            margin-bottom: 0.75rem;
            text-align: center;
            font-weight: 500;
        }
        
        .loading-progress-bar {
            height: 12px;
            background-color: #e0e0e0;
            border-radius: 6px;
            overflow: hidden;
        }
        
        .loading-progress-fill {
            height: 100%;
            background: linear-gradient(90deg, var(--accent-color), var(--success-color));
            width: 0%;
            transition: width 0.3s ease;
        }
        
        .workflow-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 1rem;
        }
        
        .workflow-progress {
            width: 100%;
            height: 8px;
            background-color: #e0e0e0;
            border-radius: 4px;
            margin-top: 1rem;
        }
        
        .workflow-progress-fill {
            height: 100%;
            background-color: var(--accent-color);
            border-radius: 4px;
            transition: width 0.3s ease;
        }
        
        .step-container {
            display: flex;
            justify-content: space-between;
            width: 100%;
        }
        
        .pareto-container {
            margin-bottom: 3rem;
        }
        
        .pareto-header {
            font-size: 1.4rem;
            font-weight: 600;
            margin-bottom: 1rem;
            color: var(--primary-color);
            border-bottom: 2px solid var(--accent-color);
            padding-bottom: 0.5rem;
        }
        
        .pareto-description {
            font-size: 1rem;
            color: var(--light-text);
            margin-bottom: 1.5rem;
        }
        
        .solution-tabs {
            margin-bottom: 1.5rem;
        }
        
        .recycled-content-table {
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 2rem;
        }
        
        .recycled-content-table th, .recycled-content-table td {
            padding: 12px;
            border: 1px solid #e0e0e0;
            text-align: center;
        }
        
        .recycled-content-table th {
            background-color: var(--primary-color);
            color: white;
        }
        
        .recycled-content-table tr:nth-child(even) {
            background-color: var(--light-bg);
        }
        
        .recycled-content-highlight {
            font-weight: 600;
            background-color: #e6f3ff;
        }
        
        .compact-metric-card {
            background-color: var(--light-bg);
            padding: 1rem;
            border-radius: 6px;
            margin: 0.75rem 0;
            font-size: 0.95rem;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
            border: 1px solid #e0e0e0;
        }
        
        .optimization-header {
            background-color: #2c3e50;
            color: white;
            padding: 1.5rem;
            border-radius: 8px;
            margin-bottom: 2rem;
            box-shadow: 0 4px 8px rgba(0,0,0,0.1);
        }
        
        .optimization-title {
            font-size: 1.75rem;
            font-weight: 600;
            margin-bottom: 0.75rem;
            color: white;
        }
        
        .method-badge {
            display: inline-block;
            padding: 0.3rem 0.6rem;
            border-radius: 4px;
            font-size: 0.85rem;
            font-weight: 600;
            margin-left: 0.75rem;
        }
        
        .ws-badge {
            background-color: var(--accent-color);
            color: white;
        }
        
        .ec-badge {
            background-color: var(--success-color);
            color: white;
        }
        
        .stNumberInput>div>div>input {
            padding: 0.5rem;
        }
        
        .stSelectbox>div>div>select {
            padding: 0.5rem;
        }
        
        .stTextInput>div>div>input {
            padding: 0.5rem;
        }
        
        .stDataFrame {
            border-radius: 6px;
            overflow: hidden;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        }
        
        .stExpander {
            border-radius: 6px;
            border: 1px solid #e0e0e0;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        }
        
        .stExpander .streamlit-expanderHeader {
            font-weight: 600;
            color: var(--primary-color);
            padding: 1rem;
        }
        
        .stTab {
            border-radius: 6px;
        }
        
        .stTab .stTab-active {
            font-weight: 600;
        }
        
        .stMarkdown h1 {
            color: var(--primary-color);
            border-bottom: 2px solid var(--accent-color);
            padding-bottom: 0.5rem;
            margin-top: 2rem;
        }
        
        .stMarkdown h2 {
            color: var(--primary-color);
            margin-top: 1.5rem;
        }
        
        .stMarkdown h3 {
            color: var(--primary-color);
            margin-top: 1.25rem;
        }
        
        .stAlert {
            border-radius: 6px;
        }
        
        .stAlert .stAlert-success {
            background-color: #e8f5e9;
            border-left: 4px solid var(--success-color);
        }
        
        .stAlert .stAlert-warning {
            background-color: #fff8e1;
            border-left: 4px solid var(--warning-color);
        }
        
        .stAlert .stAlert-error {
            background-color: #ffebee;
            border-left: 4px solid var(--danger-color);
        }
        
        .stAlert .stAlert-info {
            background-color: var(--light-bg);
            border-left: 4px solid var(--accent-color);
        }
        
        .compact-info-box {
            background-color: var(--light-bg);
            padding: 0.75rem;
            border-radius: 6px;
            border-left: 4px solid var(--accent-color);
            margin: 0.5rem 0;
            font-size: 0.85rem;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        }
        
        /* New styles for the header */
        .app-header {
            background: linear-gradient(135deg, #2c3e50 0%, #4a6491 100%);
            color: white;  /* Changed to white */
            padding: 2rem;
            border-radius: 0 0 10px 10px;
            margin-bottom: 2rem;
            box-shadow: 0 4px 12px rgba(0,0,0,0.15);
            text-align: center;
        }
        
        .app-title {
            font-size: 2.5rem;
            font-weight: 700;
            margin-bottom: 0.5rem;
            letter-spacing: 0.5px;
            text-shadow: 1px 1px 3px rgba(0,0,0,0.2);
            color: white;  /* Pure white for title */
        }
        
        .app-subtitle {
            font-size: 1.2rem;
            font-weight: 400;
            opacity: 0.9;
            margin-bottom: 0;
            color: #e0e0e0;  /* Light gray for subtitle */
        }
        
        .tradeoff-analysis {
            background-color: #f8f9fa;
            padding: 1.5rem;
            border-radius: 8px;
            border-left: 4px solid var(--accent-color);
            margin: 1.5rem 0;
        }
        
        .tradeoff-section {
            margin-bottom: 1.5rem;
        }
        
        .tradeoff-section-title {
            font-size: 1.5rem;
            font-weight: 600;
            color: var(--primary-color);
            margin-bottom: 1rem;
            padding: 0.75rem;
            background-color: var(--light-bg);
            border-radius: 6px;
            display: flex;
            align-items: center;
            gap: 0.75rem;
        }
        
        .tradeoff-section-title:before {
            content: "📊";
            font-size: 1.5rem;
        }
        
        .tradeoff-subsection-title {
            font-size: 1.2rem;
            font-weight: 600;
            color: var(--primary-color);
            margin-bottom: 0.75rem;
            padding-left: 1.5rem;
            display: flex;
            align-items: center;
            gap: 0.5rem;
        }
        
        .tradeoff-subsection-title:before {
            content: " ";
            font-size: 1.2rem;
        }
        
        .recycled-content-chart {
            background-color: white;
            padding: 1rem;
            border-radius: 8px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
            margin-bottom: 1.5rem;
        }
        
        .recycled-definition {
            font-size: 0.9rem;
            color: var(--light-text);
            margin-top: 0.5rem;
            padding: 0.75rem;
            background-color: var(--light-bg);
            border-radius: 6px;
        }
        
        .material-impact {
            font-size: 0.9rem;
            color: var(--dark-text);
            margin-top: 0.5rem;
            padding: 0.75rem;
            background-color: #fff8e1;
            border-radius: 6px;
            border-left: 4px solid var(--warning-color);
        }
    </style>
""", unsafe_allow_html=True)

# Initialize session state variables
if 'data_loaded' not in st.session_state:
    st.session_state.data_loaded = False
    st.session_state.processed_file_name = None
    st.session_state.unique_solutions = None
    st.session_state.e_constraint_solutions = None
    st.session_state.selected_grade = None
    st.session_state.chemistry_constraints = None
    st.session_state.default_chemistry = None
    st.session_state.editing_chemistry = False
    st.session_state.editing_flux_fuel = False
    st.session_state.flux_fuel_data = None
    st.session_state.electricity_ef = {
        'RE_scope2': 0.00,
        'RE_scope3': 0.31,
        'Grid_scope2': 0.727,
        'Grid_scope3': 0.31
    }
    st.session_state.usability_constraints = None
    st.session_state.editing_usability = False
    st.session_state.custom_groups = {}
    st.session_state.show_custom_group_creator = False
    st.session_state.current_step = 1
    st.session_state.selected_route = None
    
    st.session_state.LM_target = None
    st.session_state.materials = None
    st.session_state.cost = None
    st.session_state.yield_factors = None
    st.session_state.scope1 = None
    st.session_state.scope3 = None
    st.session_state.scrap_flags = None
    st.session_state.inv_min = {}
    st.session_state.inv_max = {}
    st.session_state.material_group = None
    st.session_state.group_constraints = None
    st.session_state.composition = None

# Define default chemistry constraints
DEFAULT_GRADES = {
    "304L": {
        "EAF": {
            "Mn": {"min": 1.0, "max": 1.1, "recovery": 95.0},
            "S": {"min": 0.0, "max": 0.05, "recovery": 100.0},
            "P": {"min": 0.0, "max": 0.044, "recovery": 100.0},
            "Si": {"min": 2.404, "max": 2.41, "recovery": 100.0},
            "Ni": {"min": 8.045, "max": 8.07, "recovery": 98.0},
            "Cr": {"min": 18.15, "max": 18.25, "recovery": 95.0},
            "Cu": {"min": 0.0, "max": 0.7, "recovery": 100.0},
            "Mo": {"min": 0.0, "max": 0.35, "recovery": 100.0},
            "Al": {"min": 0.0, "max": 0.05, "recovery": 100.0},
            "V": {"min": 0.0, "max": 0.05, "recovery": 100.0},
            "B": {"min": 0.0, "max": 0.0, "recovery": 100.0},
            "Nb": {"min": 0.0, "max": 0.05, "recovery": 100.0},
            "Ti": {"min": 0.0, "max": 0.0, "recovery": 100.0},
            "W": {"min": 0.0, "max": 0.0, "recovery": 100.0},
        },
        "IF": {
            "Mn": {"min": 1.0, "max": 1.1, "recovery": 100.0},
            "S": {"min": 0.0, "max": 0.05, "recovery": 100.0},
            "P": {"min": 0.0, "max": 0.044, "recovery": 100.0},
            "Si": {"min": 2.404, "max": 2.41, "recovery": 100.0},
            "Ni": {"min": 8.045, "max": 8.07, "recovery": 98.0},
            "Cr": {"min": 18.15, "max": 18.25, "recovery": 95.0},
            "Cu": {"min": 0.0, "max": 0.7, "recovery": 100.0},
            "Mo": {"min": 0.0, "max": 0.35, "recovery": 100.0},
            "Al": {"min": 0.0, "max": 0.05, "recovery": 100.0},
            "V": {"min": 0.0, "max": 0.05, "recovery": 100.0},
            "B": {"min": 0.0, "max": 0.0, "recovery": 100.0},
            "Nb": {"min": 0.0, "max": 0.05, "recovery": 100.0},
            "Ti": {"min": 0.0, "max": 0.0, "recovery": 100.0},
            "W": {"min": 0.0, "max": 0.0, "recovery": 100.0},
        }
    },
    "316L": {
        "EAF": {
            "Mn": {"min": 1.0, "max": 2.0, "recovery": 95.0},
            "S": {"min": 0.0, "max": 0.03, "recovery": 100.0},
            "P": {"min": 0.0, "max": 0.045, "recovery": 100.0},
            "Si": {"min": 0.0, "max": 1.0, "recovery": 90.0},
            "Ni": {"min": 10.0, "max": 14.0, "recovery": 98.0},
            "Cr": {"min": 16.0, "max": 18.0, "recovery": 98.0},
            "Cu": {"min": 0.0, "max": 0.7, "recovery": 100.0},
            "Mo": {"min": 2.0, "max": 3.0, "recovery": 95.0},
            "Al": {"min": 0.0, "max": 0.05, "recovery": 100.0},
            "V": {"min": 0.0, "max": 0.05, "recovery": 100.0},
            "B": {"min": 0.0, "max": 0.0, "recovery": 100.0},
            "Nb": {"min": 0.0, "max": 0.05, "recovery": 100.0},
            "Ti": {"min": 0.0, "max": 0.0, "recovery": 100.0},
            "W": {"min": 0.0, "max": 0.0, "recovery": 100.0},
        },
        "IF": {
            "Mn": {"min": 1.0, "max": 2.0, "recovery": 95.0},
            "S": {"min": 0.0, "max": 0.03, "recovery": 100.0},
            "P": {"min": 0.0, "max": 0.045, "recovery": 100.0},
            "Si": {"min": 0.0, "max": 1.0, "recovery": 90.0},
            "Ni": {"min": 10.0, "max": 14.0, "recovery": 98.0},
            "Cr": {"min": 16.0, "max": 18.0, "recovery": 98.0},
            "Cu": {"min": 0.0, "max": 0.7, "recovery": 100.0},
            "Mo": {"min": 2.0, "max": 3.0, "recovery": 95.0},
            "Al": {"min": 0.0, "max": 0.05, "recovery": 100.0},
            "V": {"min": 0.0, "max": 0.05, "recovery": 100.0},
            "B": {"min": 0.0, "max": 0.0, "recovery": 100.0},
            "Nb": {"min": 0.0, "max": 0.05, "recovery": 100.0},
            "Ti": {"min": 0.0, "max": 0.0, "recovery": 100.0},
            "W": {"min": 0.0, "max": 0.0, "recovery": 100.0},
        }
    },
    "430": {
        "EAF": {
            "Mn": {"min": 0.0, "max": 1.0, "recovery": 100.0},
            "S": {"min": 0.0, "max": 0.03, "recovery": 100.0},
            "P": {"min": 0.0, "max": 0.04, "recovery": 100.0},
            "Si": {"min": 0.0, "max": 1.0, "recovery": 90.0},
            "Ni": {"min": 0.0, "max": 0.75, "recovery": 100.0},
            "Cr": {"min": 16.0, "max": 18.0, "recovery": 95.0},
            "Cu": {"min": 0.0, "max": 0.5, "recovery": 100.0},
            "Mo": {"min": 0.0, "max": 0.75, "recovery": 100.0},
            "Al": {"min": 0.0, "max": 0.1, "recovery": 100.0},
            "V": {"min": 0.0, "max": 0.05, "recovery": 100.0},
            "B": {"min": 0.0, "max": 0.0, "recovery": 100.0},
            "Nb": {"min": 0.0, "max": 0.05, "recovery": 100.0},
            "Ti": {"min": 0.0, "max": 0.0, "recovery": 100.0},
            "W": {"min": 0.0, "max": 0.0, "recovery": 100.0},
        }
    }
}

# Default flux/fuel data
DEFAULT_FLUX_FUEL_DATA = [
    {"Material": "Lime", "Route": "EAF", "Consumption": 0.061969681, "Scope 1": 0.0, "Scope 3": 0.062},
    {"Material": "Dolomite", "Route": "EAF", "Consumption": 0.090203725, "Scope 1": 0.0, "Scope 3": 0.088},
    {"Material": "Electrode", "Route": "EAF", "Consumption": 0.003263366, "Scope 1": 3.66, "Scope 3": 0.65},
    {"Material": "Coke", "Route": "EAF", "Consumption": 0.010543205, "Scope 1": 2.98, "Scope 3": 0.03},
    {"Material": "Fluorspar", "Route": "EAF", "Consumption": 0.010205442, "Scope 1": 0.0, "Scope 3": 0.167},
    {"Material": "LDO", "Route": "EAF", "Consumption": 0.001130496, "Scope 1": 2.99, "Scope 3": 0.75},
    {"Material": "LSHS", "Route": "EAF", "Consumption": 0.002075274, "Scope 1": 3.2, "Scope 3": 0.71},
    {"Material": "Propane", "Route": "EAF", "Consumption": 0.001917105, "Scope 1": 3.01, "Scope 3": 0.35},
    {"Material": "Lime", "Route": "IF", "Consumption": 0.044278961, "Scope 1": 0.0, "Scope 3": 0.062},
    {"Material": "Dolomite", "Route": "IF", "Consumption": 0.091296898, "Scope 1": 0.0, "Scope 3": 0.088},
    {"Material": "Electrode", "Route": "IF", "Consumption": 0.000252091, "Scope 1": 3.66, "Scope 3": 0.65},
    {"Material": "Coke", "Route": "IF", "Consumption": 0.005422418, "Scope 1": 2.98, "Scope 3": 0.03},
    {"Material": "Fluorspar", "Route": "IF", "Consumption": 0.005433148, "Scope 1": 0.0, "Scope 3": 0.167},
    {"Material": "LDO", "Route": "IF", "Consumption": 0.000019748, "Scope 1": 2.99, "Scope 3": 0.75},
    {"Material": "LSHS", "Route": "IF", "Consumption": 0.001469984, "Scope 1": 3.2, "Scope 3": 0.71},
    {"Material": "Propane", "Route": "IF", "Consumption": 0.002307447, "Scope 1": 3.01, "Scope 3": 0.35}
]

# Workflow steps
WORKFLOW_STEPS = [
    {"id": 1, "label": "Upload Data", "icon": "📁"},
    {"id": 2, "label": "Set Route & Grade", "icon": "🏭"},
    {"id": 3, "label": "Set Constraints", "icon": "⚖️"},
    {"id": 4, "label": "Optimize", "icon": "⚙️"},
    {"id": 5, "label": "View Results", "icon": "📊"}
]

def show_workflow_progress():
    st.markdown('<div class="workflow-container">', unsafe_allow_html=True)
    st.markdown('<div class="step-container">', unsafe_allow_html=True)
    for step in WORKFLOW_STEPS:
        status_class = ""
        if step["id"] < st.session_state.current_step:
            status_class = "step-completed"
        elif step["id"] == st.session_state.current_step:
            status_class = "step-active"
            
        st.markdown(f"""
            <div class="workflow-step {status_class}">
                <div class="step-icon">{step["icon"]}</div>
                <div class="step-label">{step["label"]}</div>
            </div>
        """, unsafe_allow_html=True)
    
    st.markdown('</div>', unsafe_allow_html=True)
    
    progress_width = (st.session_state.current_step / len(WORKFLOW_STEPS)) * 100
    st.markdown(f"""
        <div class="workflow-progress">
            <div class="workflow-progress-fill" style="width: {progress_width}%"></div>
        </div>
    """, unsafe_allow_html=True)
    
    current_step = WORKFLOW_STEPS[st.session_state.current_step - 1]
    descriptions = {
        1: "Upload your Excel file with materials data and composition constraints",
        2: "Select the production route and steel grade",
        3: "Define material and group constraints for the optimization",
        4: "Run the optimization to find cost-emission trade-offs",
        5: "Analyze and compare the optimization results"
    }
    
    st.markdown(f'<div class="progress-text" style="text-align: center; font-style: italic;">{descriptions[st.session_state.current_step]}</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

def go_back():
    if st.session_state.current_step > 1:
        st.session_state.current_step -= 1
        st.rerun()

# File upload section
if st.session_state.current_step == 1:
    # Stylish header
    st.markdown("""
        <div class="app-header">
            <h1 class="app-title">Stainless Steel Charge-Mix Optimizer</h1>
            <p class="app-subtitle">Find optimal material combinations balancing cost and emissions</p>
        </div>
    """, unsafe_allow_html=True)
    
    show_workflow_progress()
    
    with st.expander("📁 Upload Input File", expanded=True):
        col1, col2 = st.columns(2)
        with col1:
            LM_target = st.number_input(
                "Target LM (ton)",
                min_value=1.0,
                max_value=1000.0,
                value=52.08,
                step=0.01,
                format="%.2f",
                help="Target liquid metal quantity in tons for the optimization"
            )
        
        uploaded_file = st.file_uploader(
            "Upload Excel input file", 
            type=["xlsx"],
            help="The Excel file should contain 'Materials_data' and 'Composition_constraint' sheets."
        )

        if uploaded_file:
            if not st.session_state.data_loaded or st.session_state.processed_file_name != uploaded_file.name:
                with st.spinner(f"🔍 Processing {uploaded_file.name}..."):
                    st.session_state.processed_file_name = uploaded_file.name
                    st.session_state.unique_solutions = None
                    st.session_state.e_constraint_solutions = None

                    try:
                        xls = pd.ExcelFile(uploaded_file)
                        
                        required_sheets = ["Materials_data", "Composition_constraint"]
                        missing_sheets = [sheet for sheet in required_sheets if sheet not in xls.sheet_names]
                        if missing_sheets:
                            st.error(f"❌ Missing required sheets in the Excel file: {', '.join(missing_sheets)}")
                            st.stop()
                        
                        st.session_state.LM_target = LM_target

                        materials_df = xls.parse("Materials_data")
                        materials_df.columns = materials_df.columns.str.strip()
                        
                        if materials_df['Cost'].dtype == 'object':
                            materials_df['Cost'] = materials_df['Cost'].str.replace(',', '').astype(float)
                        
                        if materials_df['Yield'].dtype == 'object':
                            materials_df['Yield'] = materials_df['Yield'].str.replace('%', '').astype(float) 
                        else:
                            materials_df['Yield'] = materials_df['Yield'] 
                        
                        materials_df = materials_df[materials_df['Cost'] > 30000].copy()
                        
                        if materials_df.empty:
                            st.error("❌ No materials remaining after filtering (cost > 30000). Please check your data.")
                            st.stop()

                        valid_materials = materials_df["Material"].tolist()
                        st.session_state.materials = valid_materials

                        composition_df_parsed = xls.parse("Composition_constraint", index_col="Material")
                        if not composition_df_parsed.index.is_unique:
                            st.warning("⚠️ Duplicate material names found in Composition sheet. Using first occurrence.")
                            composition_df_parsed = composition_df_parsed[~composition_df_parsed.index.duplicated(keep='first')]
                        composition_df_filtered = composition_df_parsed[composition_df_parsed.index.isin(valid_materials)]
                        st.session_state.composition = composition_df_filtered.T.to_dict()

                        st.session_state.cost = dict(zip(materials_df["Material"], materials_df["Cost"]))
                        st.session_state.yield_factors = dict(zip(materials_df["Material"], materials_df["Yield"]))
                        st.session_state.scope1 = dict(zip(materials_df["Material"], materials_df["Scope 1"]))
                        st.session_state.scope3 = dict(zip(materials_df["Material"], materials_df["Scope 3"]))
                        st.session_state.scrap_flags = dict(zip(materials_df["Material"], materials_df["Scrap"]))
                        st.session_state.inv_min = {mat: 0.0 for mat in valid_materials}
                        st.session_state.inv_max = {mat: 0.0 for mat in valid_materials}
                        st.session_state.material_group = dict(zip(materials_df["Material"], materials_df["Group"]))

                        st.session_state.group_constraints = (
                            materials_df.groupby("Group")
                            .agg({"Group Min": "first", "Group Max": "first"})
                            .to_dict(orient="index")
                        )
                        
                        st.session_state.usability_constraints = {
                            mat: {"min": 0.0, "max": 0.0} for mat in valid_materials
                        }
                        
                        st.session_state.custom_groups = {}
                        
                        if "Flux_fuel" in xls.sheet_names:
                            flux_fuel_df = xls.parse("Flux_fuel")
                            flux_fuel_df.columns = flux_fuel_df.columns.str.strip()
                            flux_fuel_df = flux_fuel_df.dropna(how='all')
                            
                            numeric_cols = ["Consumption", "Scope 1", "Scope 3"]
                            for col in numeric_cols:
                                if col in flux_fuel_df.columns:
                                    if flux_fuel_df[col].dtype == 'object':
                                        flux_fuel_df[col] = flux_fuel_df[col].str.replace(',', '').astype(float)
                                    else:
                                        flux_fuel_df[col] = flux_fuel_df[col].astype(float)
                            
                            for col in numeric_cols:
                                if col in flux_fuel_df.columns:
                                    flux_fuel_df[col] = flux_fuel_df[col].fillna(0)
                            
                            if "Consumption" in flux_fuel_df.columns:
                                flux_fuel_df["Consumption"] = flux_fuel_df["Consumption"] / 1000
                            
                            if "Route" in flux_fuel_df.columns:
                                flux_fuel_df["Route"] = flux_fuel_df["Route"].str.strip().str.upper()
                            else:
                                flux_fuel_df["Route"] = "EAF"
                            
                            st.session_state.flux_fuel_data = flux_fuel_df.to_dict(orient='records')
                        
                        st.session_state.data_loaded = True
                        st.success(f"✅ Successfully loaded {len(valid_materials)} materials and constraints!")
                        st.session_state.current_step = 2
                        st.rerun()
                        
                    except Exception as e:
                        st.error(f"❌ Error processing file: {str(e)}")
                        st.stop()

    # Navigation buttons
    col1, col2 = st.columns([1, 1])
    with col1:
        st.button("⬅️ Back", disabled=True, help="No previous step")
    with col2:
        if st.session_state.data_loaded:
            if st.button("➡️ Next: Set Route & Grade", type="primary"):
                st.session_state.current_step = 2
                st.rerun()
        else:
            st.button("➡️ Next: Set Route & Grade", disabled=True, help="Please upload and process a file first")

# Route and grade selection section
elif st.session_state.current_step == 2:
    # Stylish header
    st.markdown("""
        <div class="app-header">
            <h1 class="app-title">Stainless Steel Charge-Mix Optimizer</h1>
            <p class="app-subtitle">Find optimal material combinations balancing cost and emissions</p>
        </div>
    """, unsafe_allow_html=True)
    
    show_workflow_progress()
    
    if st.session_state.data_loaded:
        with st.expander("🏭 Route & Grade Selection", expanded=True):
            selected_route = st.selectbox(
                "Select Production Route",
                options=["EAF", "IF"],
                index=0,
                key="route_select"
            )
            st.session_state.selected_route = selected_route
            
            available_grades = [grade for grade in DEFAULT_GRADES.keys() if selected_route in DEFAULT_GRADES[grade]]
            
            if not available_grades:
                st.error("No grades available for the selected route")
                st.stop()
            
            selected_grade = st.selectbox(
                "Select Steel Grade",
                options=available_grades,
                index=0,
                key="grade_select"
            )
            
            st.session_state.default_chemistry = DEFAULT_GRADES[selected_grade][selected_route]
            
            st.markdown("### Chemistry & Recovery Constraints")
            
            header_col1, header_col2 = st.columns([3, 1])
            with header_col1:
                st.markdown(f"Configure the chemical composition requirements for {selected_grade} grade ({selected_route} route):")
            with header_col2:
                if st.session_state.editing_chemistry:
                    if st.button("🔒 Lock Values", key="lock_button", help="Lock current values and proceed", type="primary"):
                        st.session_state.editing_chemistry = False
                        st.rerun()
                    if st.button("🔄 Reset to Defaults", key="reset_button", help="Reset all values to grade defaults"):
                        for element in st.session_state.default_chemistry:
                            st.session_state[f"min_{element}"] = st.session_state.default_chemistry[element]["min"]
                            st.session_state[f"max_{element}"] = st.session_state.default_chemistry[element]["max"]
                            st.session_state[f"recovery_{element}"] = st.session_state.default_chemistry[element]["recovery"]
                            st.session_state[f"min_input_{element}"] = st.session_state.default_chemistry[element]["min"]
                            st.session_state[f"max_input_{element}"] = st.session_state.default_chemistry[element]["max"]
                            st.session_state[f"recovery_input_{element}"] = st.session_state.default_chemistry[element]["recovery"]
                        st.rerun()
                else:
                    if st.button("✏️ Edit Values", key="edit_button", help="Edit chemistry constraints", type="primary"):
                        st.session_state.editing_chemistry = True
                        st.rerun()
            
            st.markdown("""
            <div class="chem-row">
                <div class="chem-col"><strong>Element</strong></div>
                <div class="chem-col"><strong>Minimum (%)</strong></div>
                <div class="chem-col"><strong>Maximum (%)</strong></div>
                <div class="chem-col"><strong>Aim Recovery (%)</strong></div>
            </div>
            """, unsafe_allow_html=True)
            
            chemistry_constraints = {}
            changed_elements = []
            
            for element, default_values in st.session_state.default_chemistry.items():
                if f"min_{element}" not in st.session_state:
                    st.session_state[f"min_{element}"] = default_values["min"]
                if f"max_{element}" not in st.session_state:
                    st.session_state[f"max_{element}"] = default_values["max"]
                if f"recovery_{element}" not in st.session_state:
                    st.session_state[f"recovery_{element}"] = default_values["recovery"]
                
                cols = st.columns(4)
                with cols[0]:
                    st.markdown(f'<div class="element-name">{element}</div>', unsafe_allow_html=True)
                
                with cols[1]:
                    min_val = st.number_input(
                        f"Min {element}",
                        min_value=0.0,
                        max_value=100.0,
                        value=float(st.session_state[f"min_{element}"]),
                        step=0.001,
                        format="%.3f",
                        key=f"min_input_{element}",
                        disabled=not st.session_state.editing_chemistry,
                        label_visibility="collapsed"
                    )
                    st.session_state[f"min_{element}"] = min_val
                    default_min = default_values["min"]
                    if abs(min_val - default_min) > 0.0001:
                        changed_elements.append(element)
                        st.markdown(f'<div class="default-value changed-value">Default: {default_min:.3f}</div>', unsafe_allow_html=True)
                    else:
                        st.markdown(f'<div class="default-value">Default: {default_min:.3f}</div>', unsafe_allow_html=True)
                
                with cols[2]:
                    max_val = st.number_input(
                        f"Max {element}",
                        min_value=0.0,
                        max_value=100.0,
                        value=float(st.session_state[f"max_{element}"]),
                        step=0.001,
                        format="%.3f",
                        key=f"max_input_{element}",
                        disabled=not st.session_state.editing_chemistry,
                        label_visibility="collapsed"
                    )
                    st.session_state[f"max_{element}"] = max_val
                    default_max = default_values["max"]
                    if abs(max_val - default_max) > 0.0001:
                        if element not in changed_elements:
                            changed_elements.append(element)
                        st.markdown(f'<div class="default-value changed-value">Default: {default_max:.3f}</div>', unsafe_allow_html=True)
                    else:
                        st.markdown(f'<div class="default-value">Default: {default_max:.3f}</div>', unsafe_allow_html=True)
                
                with cols[3]:
                    recovery = st.number_input(
                        f"Recovery {element}",
                        min_value=0.0,
                        max_value=100.0,
                        value=float(st.session_state[f"recovery_{element}"]),
                        step=0.1,
                        format="%.1f",
                        key=f"recovery_input_{element}",
                        disabled=not st.session_state.editing_chemistry,
                        label_visibility="collapsed"
                    )
                    st.session_state[f"recovery_{element}"] = recovery
                    default_recovery = default_values["recovery"]
                    if abs(recovery - default_recovery) > 0.1:
                        if element not in changed_elements:
                            changed_elements.append(element)
                        st.markdown(f'<div class="default-value changed-value">Default: {default_recovery:.1f}</div>', unsafe_allow_html=True)
                    else:
                        st.markdown(f'<div class="default-value">Default: {default_recovery:.1f}</div>', unsafe_allow_html=True)
                
                chemistry_constraints[element] = {
                    "min": min_val,
                    "max": max_val,
                    "recovery": recovery
                }
            
            if changed_elements:
                st.markdown(f"""
                    <div class="info-box">
                        <strong>Changed Elements:</strong> {', '.join(changed_elements)}<br>
                        Values in <span class="changed-value">red</span> show where you've modified the default values.
                    </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown("""
                    <div class="info-box">
                        All elements are using their default values for this grade and route.
                    </div>
                """, unsafe_allow_html=True)
            
            if not st.session_state.editing_chemistry:
                if st.button("💾 Save Chemistry Constraints", type="primary", help="Save the current chemistry constraints"):
                    st.session_state.chemistry_constraints = chemistry_constraints
                    st.session_state.selected_grade = selected_grade
                    st.session_state.editing_chemistry = False
                    st.success("✅ Chemistry constraints saved! You can now proceed to set usability constraints.")
                    st.session_state.current_step = 3
                    st.rerun()

    # Navigation buttons
    col1, col2 = st.columns([1, 1])
    with col1:
        st.button("⬅️ Back", key="back_to_upload", type="secondary")
    with col2:
        if st.session_state.chemistry_constraints is not None:
            if st.button("➡️ Next: Set Constraints", type="primary"):
                st.session_state.current_step = 3
                st.rerun()
        else:
            st.button("➡️ Next: Set Constraints", disabled=True, help="Please save chemistry constraints first")

# Usability Constraints Section
elif st.session_state.current_step == 3:
    # Stylish header
    st.markdown("""
        <div class="app-header">
            <h1 class="app-title">Stainless Steel Charge-Mix Optimizer</h1>
            <p class="app-subtitle">Find optimal material combinations balancing cost and emissions</p>
        </div>
    """, unsafe_allow_html=True)
    
    show_workflow_progress()
    
    if st.session_state.data_loaded and st.session_state.chemistry_constraints is not None:
        st.markdown("## 📊 Process Constraints")
        st.markdown(f"""
            <div class="info-box">
                <strong>Current Route:</strong> {st.session_state.selected_route}<br>
                <strong>Current Grade:</strong> {st.session_state.selected_grade}<br>
                <br>
                <strong>Set constraints on specific materials and groups:</strong><br>
                Define minimum and maximum usage limits for individual materials and groups in the charge mix.
                All values are in tons (not percentages).<br>
                <strong>Note:</strong> When both min and max are 0, the material/group is unconstrained.
            </div>
        """, unsafe_allow_html=True)
        
        if st.session_state.editing_usability:
            if st.button("🔒 Lock Values", key="lock_usability_button", help="Lock current values and proceed", type="primary"):
                st.session_state.editing_usability = False
                st.rerun()
            if st.button("🔄 Reset Values", key="reset_usability_button", help="Reset all values to defaults"):
                st.session_state.usability_constraints = {
                    mat: {"min": 0.0, "max": 0.0} for mat in st.session_state.materials
                }
                st.session_state.custom_groups = {}
                st.rerun()
        else:
            if st.button("✏️ Edit Values", key="edit_usability_button", help="Edit material constraints", type="primary"):
                st.session_state.editing_usability = True
                st.rerun()
        
        if st.session_state.editing_usability:
            search_term = st.text_input("🔍 Search Materials", "")
            
            group_materials = {}
            for mat in st.session_state.materials:
                group = st.session_state.material_group.get(mat, "Ungrouped")
                if group not in group_materials:
                    group_materials[group] = []
                group_materials[group].append(mat)
            
            with st.expander("➕ Create Custom Group", expanded=False):
                if st.button("Create New Custom Group", key="new_custom_group"):
                    st.session_state.show_custom_group_creator = True
                
                if st.session_state.show_custom_group_creator:
                    new_group_name = st.text_input("Custom Group Name", value="NewGroup", key="new_group_name")
                    
                    st.markdown("Select materials for this group:")
                    selected_materials = []
                    
                    with st.container():
                        st.markdown('<div class="material-selector">', unsafe_allow_html=True)
                        for mat in st.session_state.materials:
                            selected = st.checkbox(
                                mat, 
                                key=f"group_select_{mat}",
                                value=mat in selected_materials
                            )
                            if selected:
                                selected_materials.append(mat)
                        st.markdown('</div>', unsafe_allow_html=True)
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        group_min = st.number_input(
                            "Minimum Amount (ton)", 
                            min_value=0.0,
                            max_value=st.session_state.LM_target,
                            value=0.0,
                            step=0.0001,
                            format="%.4f",
                            key="new_group_min"
                        )
                    with col2:
                        group_max = st.number_input(
                            "Maximum Amount (ton)", 
                            min_value=0.0,
                            max_value=st.session_state.LM_target,
                            value=0.0,
                            step=0.0001,
                            format="%.4f",
                            key="new_group_max"
                        )
                    
                    validation_errors = []
                    if group_min > 0 and group_max > 0 and group_min > group_max:
                        validation_errors.append("Group minimum cannot be greater than group maximum")
                    
                    for mat in selected_materials:
                        mat_min = st.session_state.usability_constraints[mat]["min"]
                        mat_max = st.session_state.usability_constraints[mat]["max"]
                        
                        if group_max > 0 and mat_min > 0 and group_max < mat_min:
                            validation_errors.append(f"Group maximum ({group_max:.4f} ton) is less than {mat}'s minimum ({mat_min:.4f} ton)")
                        if group_min > 0 and mat_max > 0 and group_min > mat_max:
                            validation_errors.append(f"Group minimum ({group_min:.4f} ton) is greater than {mat}'s maximum ({mat_max:.4f} ton)")
                    
                    if validation_errors:
                        for error in validation_errors:
                            st.error(error)
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button("💾 Save Custom Group", key="save_custom_group") and not validation_errors:
                            if new_group_name and selected_materials:
                                st.session_state.custom_groups[new_group_name] = {
                                    "materials": selected_materials,
                                    "min": group_min,
                                    "max": group_max
                                }
                                st.success(f"Custom group '{new_group_name}' saved with {len(selected_materials)} materials!")
                                st.session_state.show_custom_group_creator = False
                                st.rerun()
                            else:
                                st.warning("Please provide a group name and select at least one material")
                    with col2:
                        if st.button("❌ Cancel", key="cancel_custom_group"):
                            st.session_state.show_custom_group_creator = False
                            st.rerun()
            
            if st.session_state.custom_groups:
                st.markdown("### Custom Groups")
                for group_name, group_data in st.session_state.custom_groups.items():
                    with st.container():
                        st.markdown(f'<div class="custom-group-container">', unsafe_allow_html=True)
                        st.markdown(f'<div class="custom-group-header">Group: {group_name}</div>', unsafe_allow_html=True)
                        
                        cols = st.columns([3, 1, 1, 1])
                        with cols[0]:
                            st.markdown(f"<div class='constraint-summary'>Materials: {', '.join(group_data['materials'])}</div>", unsafe_allow_html=True)
                        
                        validation_errors = []
                        for mat in group_data["materials"]:
                            mat_min = st.session_state.usability_constraints[mat]["min"]
                            mat_max = st.session_state.usability_constraints[mat]["max"]
                            
                            if group_data["max"] > 0 and mat_min > 0 and group_data["max"] < mat_min:
                                validation_errors.append(f"Group maximum ({group_data['max']:.4f} ton) is less than {mat}'s minimum ({mat_min:.4f} ton)")
                            if group_data["min"] > 0 and mat_max > 0 and group_data["min"] > mat_max:
                                validation_errors.append(f"Group minimum ({group_data['min']:.4f} ton) is greater than {mat}'s maximum ({mat_max:.4f} ton)")
                        
                        with cols[1]:
                            new_min = st.number_input(
                                f"Minimum Amount (ton) for {group_name}",
                                min_value=0.0,
                                max_value=st.session_state.LM_target,
                                value=float(group_data['min']),
                                step=0.0001,
                                format="%.4f",
                                key=f"custom_min_{group_name}"
                            )
                            group_data['min'] = new_min
                        with cols[2]:
                            new_max = st.number_input(
                                f"Maximum Amount (ton) for {group_name}",
                                min_value=0.0,
                                max_value=st.session_state.LM_target,
                                value=float(group_data['max']),
                                step=0.0001,
                                format="%.4f",
                                key=f"custom_max_{group_name}"
                            )
                            group_data['max'] = new_max
                        with cols[3]:
                            if st.button("🗑️", key=f"delete_{group_name}"):
                                del st.session_state.custom_groups[group_name]
                                st.rerun()
                        
                        if validation_errors:
                            st.markdown('<div class="constraint-error">⚠️ Constraints conflict with individual material constraints:</div>', unsafe_allow_html=True)
                            for error in validation_errors:
                                st.markdown(f'<div class="constraint-error">{error}</div>', unsafe_allow_html=True)
                        
                        st.markdown('</div>', unsafe_allow_html=True)
            
            st.markdown("### Predefined Groups from Input File")
            for group, materials in group_materials.items():
                filtered_materials = [mat for mat in materials if search_term.lower() in mat.lower()]
                
                if filtered_materials:
                    with st.expander(f"Group: {group}", expanded=False):
                        st.markdown(f'<div class="group-container">', unsafe_allow_html=True)
                        
                        group_min = st.session_state.group_constraints.get(group, {}).get("Group Min", 0)
                        group_max = st.session_state.group_constraints.get(group, {}).get("Group Max", 0)
                        
                        validation_errors = []
                        for mat in filtered_materials:
                            mat_min = st.session_state.usability_constraints[mat]["min"]
                            mat_max = st.session_state.usability_constraints[mat]["max"]
                            
                            if group_max > 0 and mat_min > 0 and group_max < mat_min:
                                validation_errors.append(f"Group maximum ({group_max:.4f} ton) is less than {mat}'s minimum ({mat_min:.4f} ton)")
                            if group_min > 0 and mat_max > 0 and group_min > mat_max:
                                validation_errors.append(f"Group minimum ({group_min:.4f} ton) is greater than {mat}'s maximum ({mat_max:.4f} ton)")
                        
                        cols = st.columns([3, 1, 1])
                        with cols[0]:
                            st.markdown(f"<div class='constraint-summary'>Materials: {', '.join(filtered_materials)}</div>", unsafe_allow_html=True)
                        with cols[1]:
                            new_min = st.number_input(
                                f"Minimum Amount (ton) for {group}",
                                min_value=0.0,
                                max_value=st.session_state.LM_target,
                                value=float(group_min),
                                step=0.0001,
                                format="%.4f",
                                key=f"group_min_{group}"
                            )
                            st.session_state.group_constraints[group]["Group Min"] = new_min
                        with cols[2]:
                            new_max = st.number_input(
                                f"Maximum Amount (ton) for {group}",
                                min_value=0.0,
                                max_value=st.session_state.LM_target,
                                value=float(group_max),
                                step=0.0001,
                                format="%.4f",
                                key=f"group_max_{group}"
                            )
                            st.session_state.group_constraints[group]["Group Max"] = new_max
                        
                        if validation_errors:
                            st.markdown('<div class="constraint-error">⚠️ Constraints conflict with individual material constraints:</div>', unsafe_allow_html=True)
                            for error in validation_errors:
                                st.markdown(f'<div class="constraint-error">{error}</div>', unsafe_allow_html=True)
                        
                        for mat in filtered_materials:
                            st.markdown(f'<div class="material-constraint-row">', unsafe_allow_html=True)
                            cols = st.columns([3, 2, 2, 1])
                            
                            with cols[0]:
                                st.markdown(f'<div class="element-name">{mat}</div>', unsafe_allow_html=True)
                            
                            with cols[1]:
                                st.markdown("Minimum Amount")
                                usability_min = st.number_input(
                                    f"Minimum Amount {mat}",
                                    min_value=0.0,
                                    max_value=st.session_state.LM_target,
                                    value=float(st.session_state.usability_constraints[mat]["min"]),
                                    step=0.0001,
                                    format="%.4f",
                                    key=f"usability_min_{mat}",
                                    label_visibility="collapsed"
                                )
                                st.session_state.usability_constraints[mat]["min"] = usability_min
                            
                            with cols[2]:
                                st.markdown("Maximum Amount")
                                usability_max = st.number_input(
                                    f"Maximum Amount {mat}",
                                    min_value=0.0,
                                    max_value=st.session_state.LM_target,
                                    value=float(st.session_state.usability_constraints[mat]["max"]),
                                    step=0.0001,
                                    format="%.4f",
                                    key=f"usability_max_{mat}",
                                    label_visibility="collapsed"
                                )
                                st.session_state.usability_constraints[mat]["max"] = usability_max
                            
                            with cols[3]:
                                if st.button("🗑️", key=f"reset_usability_{mat}"):
                                    st.session_state.usability_constraints[mat] = {"min": 0.0, "max": 0.0}
                                    st.rerun()
                            
                            st.markdown('</div>', unsafe_allow_html=True)
                        
                        st.markdown('</div>', unsafe_allow_html=True)
        else:
            constrained_materials = [
                mat for mat, constraints in st.session_state.usability_constraints.items() 
                if constraints["min"] > 0 or constraints["max"] > 0
            ]
            
            constrained_groups = []
            
            for group, constraints in st.session_state.group_constraints.items():
                if constraints["Group Min"] > 0 or constraints["Group Max"] > 0:
                    constrained_groups.append({
                        "Type": "Predefined",
                        "Group": group,
                        "Minimum Amount (ton)": constraints["Group Min"],
                        "Maximum Amount (ton)": constraints["Group Max"]
                    })
            
            for group_name, group_data in st.session_state.custom_groups.items():
                if group_data["min"] > 0 or group_data["max"] > 0:
                    constrained_groups.append({
                        "Type": "Custom",
                        "Group": group_name,
                        "Minimum Amount (ton)": group_data["min"],
                        "Maximum Amount (ton)": group_data["max"],
                        "Materials": ", ".join(group_data["materials"])
                    })
            
            if constrained_materials or constrained_groups:
                st.markdown("### Current Constraints Summary")
                
                if constrained_materials:
                    st.markdown("#### Material-Specific Constraints")
                    material_data = []
                    for mat in constrained_materials:
                        constraints = st.session_state.usability_constraints[mat]
                        material_data.append({
                            "Material": mat,
                            "Group": st.session_state.material_group.get(mat, "Ungrouped"),
                            "Minimum Amount (ton)": constraints["min"],
                            "Maximum Amount (ton)": constraints["max"]
                        })
                    
                    st.dataframe(
                        pd.DataFrame(material_data).sort_values("Group"),
                        use_container_width=True
                    )
                
                if constrained_groups:
                    st.markdown("#### Group-Level Constraints")
                    group_df = pd.DataFrame(constrained_groups)
                    
                    if "Materials" in group_df.columns:
                        group_df = group_df[["Type", "Group", "Materials", "Minimum Amount (ton)", "Maximum Amount (ton)"]]
                    else:
                        group_df = group_df[["Type", "Group", "Minimum Amount (ton)", "Maximum Amount (ton)"]]
                    
                    st.dataframe(
                        group_df,
                        use_container_width=True
                    )
            else:
                st.markdown("""
                    <div class="info-box">
                        No material-specific or group-level constraints have been set. All materials are unconstrained.
                    </div>
                """, unsafe_allow_html=True)

    # Navigation buttons
    col1, col2 = st.columns([1, 1])
    with col1:
        if st.button("⬅️ Back", key="back_to_chemistry", type="secondary"):
            st.session_state.current_step = 2
            st.rerun()
    with col2:
        if not st.session_state.editing_usability:
            if st.button("➡️ Next: Optimize", type="primary"):
                st.session_state.current_step = 4
                st.rerun()
        else:
            st.button("➡️ Next: Optimize", disabled=True, help="Please lock usability constraints first")

# Optimization section
elif st.session_state.current_step == 4:
    # Stylish header
    st.markdown("""
        <div class="app-header">
            <h1 class="app-title">Stainless Steel Charge-Mix Optimizer</h1>
            <p class="app-subtitle">Find optimal material combinations balancing cost and emissions</p>
        </div>
    """, unsafe_allow_html=True)
    
    show_workflow_progress()
    
    if st.session_state.data_loaded and st.session_state.chemistry_constraints is not None:
        current_materials = st.session_state.materials
        current_cost = st.session_state.cost
        current_yield_factors = st.session_state.yield_factors
        current_scope1 = st.session_state.scope1
        current_scope3 = st.session_state.scope3
        current_composition = st.session_state.composition
        current_chemistry_constraints = st.session_state.chemistry_constraints
        current_inv_min = st.session_state.inv_min
        current_inv_max = st.session_state.inv_max
        current_material_group = st.session_state.material_group
        current_group_constraints = st.session_state.group_constraints
        current_scrap_flags = st.session_state.scrap_flags
        current_usability_constraints = st.session_state.usability_constraints
        current_custom_groups = st.session_state.custom_groups
        current_LM_target = st.session_state.LM_target

        st.markdown(f"""
            <div class="info-box">
                <strong>Current Route:</strong> {st.session_state.selected_route}<br>
                <strong>Current Grade:</strong> {st.session_state.selected_grade}<br>
                <strong>Elements Configured:</strong> {', '.join(current_chemistry_constraints.keys())}
            </div>
        """, unsafe_allow_html=True)

        with st.expander("⚙️ Optimization Settings", expanded=True):
            col1, col2 = st.columns([1, 3])
            with col1:
                if st.button("🚀 Run Optimization", type="primary", help="Run the optimization with current settings"):
                    with st.spinner("🔄 Running optimization scenarios... This might take a moment."):
                        try:
                            # Run weighted sum method
                            weight_combinations = [(w1, 1 - w1) for w1 in [i / 100 for i in range(101)]]
                            solutions = []
                            
                            # Update progress for weighted sum method
                            progress_text = st.empty()
                            progress_bar = st.progress(0)
                            
                            for i, (w1, w2) in enumerate(weight_combinations):
                                progress_text.text(f"Running weighted sum optimization ({i+1}/{len(weight_combinations)}) - {w1:.2f}:{w2:.2f}")
                                progress_bar.progress((i + 1) / len(weight_combinations))
                                
                                mix, cost_val, emission_val = optimize_mixture_core(
                                    w1, w2, current_materials, current_cost, current_yield_factors,
                                    current_scope1, current_scope3, current_composition, current_chemistry_constraints,
                                    current_inv_min, current_inv_max, current_material_group, current_group_constraints,
                                    current_scrap_flags, current_usability_constraints, current_custom_groups, current_LM_target
                                )
                                solutions.append((mix, cost_val, emission_val))
                            
                            def round_mix(mix, decimals=4):
                                return {k: round(v, decimals) for k, v in mix.items()}

                            unique_sols_list = []
                            seen_mixes = set()

                            for i, (mix, cost_val, emission_val) in enumerate(solutions):
                                if not mix:
                                    continue
                                rounded_mix_tuple = frozenset(round_mix(mix).items())
                                if rounded_mix_tuple not in seen_mixes:
                                    seen_mixes.add(rounded_mix_tuple)
                                    w1_s, w2_s = weight_combinations[i]
                                    
                                    total_input_kg = sum(mix.values())
                                    
                                    scrap_input_kg = sum(mix[mat] for mat in mix if current_scrap_flags.get(mat, 0) == 1)
                                    scrap_pct_s = (scrap_input_kg / total_input_kg) * 100 if total_input_kg > 0 else 0
                                    
                                    revert_input_kg = sum(mix[mat] for mat in mix if current_material_group.get(mat) == "REVERT")
                                    revert_pct = (revert_input_kg / total_input_kg) * 100 if total_input_kg > 0 else 0
                                    
                                    post_consumer_pct = scrap_pct_s - revert_pct
                                    
                                    yield_pct = (1000 / total_input_kg) * 100 if total_input_kg > 0 else 0
                                    
                                    unique_sols_list.append({
                                        "mix": mix, 
                                        "w1": w1_s, 
                                        "w2": w2_s,
                                        "Cost (INR/ton)": cost_val,
                                        "Emissions (tCO₂/tcs)": emission_val,
                                        "Scrap %": scrap_pct_s,
                                        "Pre-Consumer %": revert_pct,
                                        "Post-Consumer %": post_consumer_pct,
                                        "Recycle Content %": scrap_pct_s,
                                        "Yield %": yield_pct,
                                        "Total Input (kg)": total_input_kg,
                                        "Total Input (ton)": total_input_kg / 1000,
                                        "Method": "Weighted Sum"
                                    })
                            
                            st.session_state.unique_solutions = unique_sols_list
                            
                            # Update progress for ε-constraint method
                            progress_text.text("Starting adaptive ε-constraint optimization")
                            progress_bar.progress(0)
                            
                            cost_opt_mix, cost_opt_cost, cost_opt_emission = optimize_mixture_core(
                                1, 0, current_materials, current_cost, current_yield_factors,
                                current_scope1, current_scope3, current_composition, current_chemistry_constraints,
                                current_inv_min, current_inv_max, current_material_group, current_group_constraints,
                                current_scrap_flags, current_usability_constraints, current_custom_groups, current_LM_target
                            )
                            
                            emission_opt_mix, emission_opt_cost, emission_opt_emission = optimize_mixture_core(
                                0, 1, current_materials, current_cost, current_yield_factors,
                                current_scope1, current_scope3, current_composition, current_chemistry_constraints,
                                current_inv_min, current_inv_max, current_material_group, current_group_constraints,
                                current_scrap_flags, current_usability_constraints, current_custom_groups, current_LM_target
                            )
                            
                            min_emission = min(emission_opt_emission, cost_opt_emission)
                            max_emission = max(emission_opt_emission, cost_opt_emission)
                            
                            e_constraint_solutions = adaptive_e_constraint_method(min_emission, max_emission)
                            
                            # Update progress during ε-constraint method
                            progress_text.text(f"Processing {len(e_constraint_solutions)} ε-constraint solutions")
                            progress_bar.progress(0.75)
                            
                            e_constraint_results = []
                            for i, (mix, cost_val, emission_val) in enumerate(e_constraint_solutions):
                                if mix:
                                    total_input_kg = sum(mix.values())
                                    
                                    scrap_input_kg = sum(mix[mat] for mat in mix if current_scrap_flags.get(mat, 0) == 1)
                                    scrap_pct_s = (scrap_input_kg / total_input_kg) * 100 if total_input_kg > 0 else 0
                                    
                                    revert_input_kg = sum(mix[mat] for mat in mix if current_material_group.get(mat) == "REVERT")
                                    revert_pct = (revert_input_kg / total_input_kg) * 100 if total_input_kg > 0 else 0
                                    
                                    post_consumer_pct = scrap_pct_s - revert_pct
                                    
                                    yield_pct = (1000 / total_input_kg) * 100 if total_input_kg > 0 else 0
                                    
                                    e_constraint_results.append({
                                        "mix": mix,
                                        "Cost (INR/ton)": cost_val,
                                        "Emissions (tCO₂/tcs)": emission_val,
                                        "Scrap %": scrap_pct_s,
                                        "Pre-Consumer %": revert_pct,
                                        "Post-Consumer %": post_consumer_pct,
                                        "Recycle Content %": scrap_pct_s,
                                        "Yield %": yield_pct,
                                        "Total Input (kg)": total_input_kg,
                                        "Total Input (ton)": total_input_kg / 1000,
                                        "Method": "ε-Constraint"
                                    })
                            
                            st.session_state.e_constraint_solutions = e_constraint_results
                            
                            # Final progress update
                            progress_text.text("Optimization finished successfully")
                            progress_bar.progress(1.0)
                            
                            if not unique_sols_list and not e_constraint_results:
                                st.warning("⚠️ Optimizer ran, but no feasible solutions were found with the given data and constraints.")
                            else:
                                st.success(f"✅ Optimization complete! Found {len(unique_sols_list)} weighted sum solution(s) and {len(e_constraint_results)} ε-constraint solution(s).")
                                st.session_state.current_step = 5
                                st.rerun()
                                
                        except Exception as e:
                            st.error(f"❌ Error during optimization: {str(e)}")
            
            with col2:
                st.markdown("""
                    <div class="compact-info-box">
                        <strong>Optimization Approaches:</strong><br>
                        1. <strong>Weighted Sum Method</strong>: Explores weight combinations between cost (w1) and emissions (w2)<br>
                        2. <strong>Adaptive ε-Constraint Method</strong>: Minimizes cost subject to emission constraints (ε) with adaptive refinement
                    </div>
                    <div class="compact-info-box">
                        <strong>Key Features:</strong><br>
                        • Each solution meets all chemical composition requirements<br>
                        • Solutions must contain at least 75% scrap material (if available)<br>
                        • Inventory, group, and material-specific constraints are respected
                    </div>
                """, unsafe_allow_html=True)

    # Navigation buttons
    col1, col2 = st.columns([1, 1])
    with col1:
        if st.button("⬅️ Back", key="back_to_constraints", type="secondary"):
            st.session_state.current_step = 3
            st.rerun()
    with col2:
        if st.session_state.unique_solutions is not None or st.session_state.e_constraint_solutions is not None:
            if st.button("➡️ Next: View Results", type="primary"):
                st.session_state.current_step = 5
                st.rerun()
        else:
            st.button("➡️ Next: View Results", disabled=True, help="Please run optimization first")

# Display results if available
elif st.session_state.current_step == 5:
    # Stylish header
    st.markdown("""
        <div class="app-header">
            <h1 class="app-title">Stainless Steel Charge-Mix Optimizer</h1>
            <p class="app-subtitle">Find optimal material combinations balancing cost and emissions</p>
        </div>
    """, unsafe_allow_html=True)
    
    show_workflow_progress()
    
    if st.session_state.unique_solutions is not None or st.session_state.e_constraint_solutions is not None:
        all_solutions = []
        
        if st.session_state.unique_solutions:
            for sol in st.session_state.unique_solutions:
                sol_copy = sol.copy()
                all_solutions.append(sol_copy)
        
        if st.session_state.e_constraint_solutions:
            for sol in st.session_state.e_constraint_solutions:
                sol_copy = sol.copy()
                all_solutions.append(sol_copy)
        
        if not all_solutions:
            st.warning("No solutions found with the current constraints.")
            st.stop()

        all_solutions_sorted = sorted(all_solutions, key=lambda x: x["Emissions (tCO₂/tcs)"])
        
        min_cost = min(sol["Cost (INR/ton)"] for sol in all_solutions_sorted)
        max_cost = max(sol["Cost (INR/ton)"] for sol in all_solutions_sorted)
        min_emissions = min(sol["Emissions (tCO₂/tcs)"] for sol in all_solutions_sorted)
        max_emissions = max(sol["Emissions (tCO₂/tcs)"] for sol in all_solutions_sorted)
        
        # Calculate recycled content statistics
        avg_scrap = sum(sol["Scrap %"] for sol in all_solutions_sorted) / len(all_solutions_sorted)
        avg_pre_consumer = sum(sol["Pre-Consumer %"] for sol in all_solutions_sorted) / len(all_solutions_sorted)
        avg_post_consumer = sum(sol["Post-Consumer %"] for sol in all_solutions_sorted) / len(all_solutions_sorted)
        avg_recycle_content = avg_scrap
        
        # Calculate yield range
        min_yield = min(sol["Yield %"] for sol in all_solutions_sorted)
        max_yield = max(sol["Yield %"] for sol in all_solutions_sorted)
        
        # Calculate scrap range
        min_scrap = min(sol["Scrap %"] for sol in all_solutions_sorted)
        max_scrap = max(sol["Scrap %"] for sol in all_solutions_sorted)
        
        # Create a compact optimization overview with highlighted route and grade
        st.markdown(f"""
            <div class="optimization-header">
                <h3 class="optimization-title">📊 Optimization Results Overview</h3>
                <p style="color: white;"><strong>Route:</strong> {st.session_state.selected_route} | <strong>Grade:</strong> {st.session_state.selected_grade}</p>
            </div>
        """, unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(f"""
                <div class="compact-metric-card">
                    <strong>Total Solutions</strong><br>
                    {len(all_solutions_sorted)}
                </div>
            """, unsafe_allow_html=True)
            
            st.markdown(f"""
                <div class="compact-metric-card">
                    <strong>Cost Range(Charge-Mix)</strong><br>
                    ₹{min_cost:,.0f} - ₹{max_cost:,.0f}/ton
                </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
                <div class="compact-metric-card">
                    <strong>Emissions Range(Charge-Mix)</strong><br>
                    {min_emissions:.4f} - {max_emissions:.4f} tCO₂/ton
                </div>
            """, unsafe_allow_html=True)
            
            st.markdown(f"""
                <div class="compact-metric-card">
                    <strong>Yield Range</strong><br>
                    {min_yield:.1f}% - {max_yield:.1f}%
                </div>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown(f"""
                <div class="compact-metric-card">
                    <strong>Average Input Weight</strong><br>
                    {sum(sol['Total Input (kg)'] for sol in all_solutions_sorted) / len(all_solutions_sorted):,.1f} kg
                </div>
            """, unsafe_allow_html=True)
            
            st.markdown(f"""
                <div class="compact-metric-card">
                    <strong>Scrap Range</strong><br>
                    {min_scrap:.1f}% - {max_scrap:.1f}%
                </div>
            """, unsafe_allow_html=True)
        
        # Recycled Content Summary - Now as Pie Chart
        st.markdown("### ♻️ Recycled Content Summary")
        
        # Create pie chart
        pie_fig = create_recycled_content_pie_chart(avg_pre_consumer, avg_post_consumer)
        st.plotly_chart(pie_fig, use_container_width=True)
        
        # Add ISO 14021 definitions
        st.markdown("""
            <div class="recycled-definition">
                <strong>Definitions (ISO 14021):</strong><br>
                • <strong>Pre-Consumer Material</strong>: Material diverted from the waste stream during a manufacturing process (e.g., revert scrap)<br>
                • <strong>Post-Consumer Material</strong>: Material generated by households or commercial, industrial and institutional facilities in their role as end-users of the product (e.g., end-of-life scrap)<br>
                • <strong>Recycled Content</strong>: Proportion, by mass, of recycled material in a product (sum of pre-consumer and post-consumer materials)
            </div>
        """, unsafe_allow_html=True)
        
        tab1, tab2 = st.tabs(["Weighted Sum Method", "Adaptive ε-Constraint Method"])
        
        with tab1:
            if st.session_state.unique_solutions:
                st.markdown('<div class="pareto-container">', unsafe_allow_html=True)
                st.markdown('<div class="pareto-header">Weighted Sum Method: Cost vs Emissions Trade-Off</div>', unsafe_allow_html=True)
                st.markdown('<div class="pareto-description">Solutions obtained by varying weights between cost and emissions objectives.</div>', unsafe_allow_html=True)
                
                plot_df = pd.DataFrame([{
                    "Solution": f"Solution {i+1}",
                    "Cost (INR/ton)": sol["Cost (INR/ton)"],
                    "Emissions (tCO₂)": sol["Emissions (tCO₂/tcs)"],
                    "Scrap %": sol["Scrap %"],
                    "Pre-Consumer %": sol["Pre-Consumer %"],
                    "Post-Consumer %": sol["Post-Consumer %"],
                    "Recycle Content %": sol["Recycle Content %"],
                    "Yield %": sol["Yield %"],
                    "w1": sol["w1"],
                    "w2": sol["w2"],
                    "Method": "Weighted Sum"
                } for i, sol in enumerate(st.session_state.unique_solutions)])
                
                fig = px.scatter(
                    plot_df,
                    x="Emissions (tCO₂)",
                    y="Cost (INR/ton)",
                    color="Recycle Content %",
                    hover_data=["Solution", "w1", "w2", "Scrap %", "Pre-Consumer %", "Post-Consumer %", "Yield %"],
                    labels={
                        "Emissions (tCO₂)": "Emissions (tCO₂/ton steel)",
                        "Cost (INR/ton)": "Cost (INR/ton)",
                        "Scrap %": "Scrap Percentage",
                        "Pre-Consumer %": "Pre-Consumer (Revert) %",
                        "Post-Consumer %": "Post-Consumer %",
                        "Recycle Content %": "Recycle Content %",
                        "Yield %": "Yield Percentage"
                    },
                    title="Weighted Sum Method: Pareto Front",
                    template="plotly_white"
                )
                
                fig.update_traces(
                    marker=dict(size=10, line=dict(width=1, color='DarkSlateGrey')),
                    selector=dict(mode='markers')
                )
                
                fig.update_layout(
                    hoverlabel=dict(
                        bgcolor="white",
                        font_size=12,
                        font_family="Arial"
                    )
                )
                
                st.plotly_chart(fig, use_container_width=True)
                
                # Trade-off analysis section
                st.markdown('<div class="tradeoff-analysis">', unsafe_allow_html=True)
                st.markdown('<div class="tradeoff-section-title">Trade-Off Analysis</div>', unsafe_allow_html=True)
                
                # Identify boundary solutions
                min_cost_solution = min(st.session_state.unique_solutions, key=lambda x: x["Cost (INR/ton)"])
                min_emission_solution = min(st.session_state.unique_solutions, key=lambda x: x["Emissions (tCO₂/tcs)"])
                
                # Calculate differences between boundary solutions
                cost_diff = min_emission_solution["Cost (INR/ton)"] - min_cost_solution["Cost (INR/ton)"]
                emission_diff = min_cost_solution["Emissions (tCO₂/tcs)"] - min_emission_solution["Emissions (tCO₂/tcs)"]
                
                st.markdown(f"""
                    <div class="tradeoff-section">
                        <p>The trade-off curve shows the relationship between cost and emissions for different charge mixes. 
                        Key observations:</p>
                        <ul>
                            <li>The <strong>cheapest solution</strong> costs ₹{min_cost_solution["Cost (INR/ton)"]:,.0f}/ton with emissions of {min_cost_solution["Emissions (tCO₂/tcs)"]:.4f} tCO₂/ton</li>
                            <li>The <strong>lowest emission solution</strong> emits {min_emission_solution["Emissions (tCO₂/tcs)"]:.4f} tCO₂/ton at a cost of ₹{min_emission_solution["Cost (INR/ton)"]:,.0f}/ton</li>
                            <li>To reduce emissions by {emission_diff:.4f} tCO₂/ton, costs increase by ₹{cost_diff:,.0f}/ton ({cost_diff/min_cost_solution["Cost (INR/ton)"]*100:.1f}%)</li>
                        </ul>
                    </div>
                """, unsafe_allow_html=True)
                
                # Identify knee point for weighted sum method
                solutions_sorted = sorted(st.session_state.unique_solutions, key=lambda x: x["Emissions (tCO₂/tcs)"])
                slopes = []
                for i in range(1, len(solutions_sorted)):
                    cost_diff = solutions_sorted[i]["Cost (INR/ton)"] - solutions_sorted[i-1]["Cost (INR/ton)"]
                    emission_diff = solutions_sorted[i]["Emissions (tCO₂/tcs)"] - solutions_sorted[i-1]["Emissions (tCO₂/tcs)"]
                    if emission_diff != 0:
                        slope = cost_diff / emission_diff
                        slopes.append((i, slope))
                
                if len(slopes) > 1:
                    max_slope_change_idx = 0
                    max_slope_change = 0
                    for i in range(1, len(slopes)):
                        slope_change = abs(slopes[i][1] - slopes[i-1][1])
                        if slope_change > max_slope_change:
                            max_slope_change = slope_change
                            max_slope_change_idx = i
                    
                    knee_point = solutions_sorted[slopes[max_slope_change_idx][0]]
                    prev_point = solutions_sorted[slopes[max_slope_change_idx][0]-1]
                    
                    # Compare material mixes to identify key differences
                    knee_materials = set(knee_point["mix"].keys())
                    prev_materials = set(prev_point["mix"].keys())
                    
                    added_materials = knee_materials - prev_materials
                    removed_materials = prev_materials - knee_materials
                    
                    st.markdown(f"""
                        <div class="tradeoff-section">
                            <div class="tradeoff-subsection-title">🔍 Knee Point Analysis</div>
                            <p>The knee point represents the optimal trade-off between cost and emissions:</p>
                            <ul>
                                <li><strong>Cost:</strong> ₹{knee_point["Cost (INR/ton)"]:,.0f}/ton</li>
                                <li><strong>Emissions:</strong> {knee_point["Emissions (tCO₂/tcs)"]:.4f} tCO₂/ton</li>
                                <li><strong>Scrap Content:</strong> {knee_point["Scrap %"]:.1f}%</li>
                            </ul>
                            <p>At this point, the cost increases by ₹{(knee_point["Cost (INR/ton)"] - prev_point["Cost (INR/ton)"]):,.0f}/ton 
                            for an emission reduction of {(prev_point["Emissions (tCO₂/tcs)"] - knee_point["Emissions (tCO₂/tcs)"]):.4f} tCO₂/ton.</p>
                        </div>
                    """, unsafe_allow_html=True)
                    
                    if added_materials or removed_materials:
                        st.markdown('<div class="tradeoff-section">', unsafe_allow_html=True)
                        st.markdown('<div class="tradeoff-subsection-title">🔧 Key Material Changes at Knee Point</div>', unsafe_allow_html=True)
                        if added_materials:
                            st.markdown(f'<p>Materials added: {", ".join(added_materials)}</p>', unsafe_allow_html=True)
                        if removed_materials:
                            st.markdown(f'<p>Materials removed: {", ".join(removed_materials)}</p>', unsafe_allow_html=True)
                        st.markdown('<p>This suggests that higher-cost, lower-emission materials are being substituted at this point.</p>', unsafe_allow_html=True)
                        st.markdown('</div>', unsafe_allow_html=True)
                
                st.markdown('</div>', unsafe_allow_html=True)  # Close tradeoff-analysis
                
                st.markdown('</div>', unsafe_allow_html=True)  # Close pareto-container
                
                # Create Excel download with detailed charge mixes
                excel_file = create_excel_download(st.session_state.unique_solutions)
                st.download_button(
                    label="📥 Download Weighted Sum Solutions (Excel)",
                    data=excel_file,
                    file_name="weighted_sum_solutions.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                # Create heat template download
                composition_df = pd.DataFrame(st.session_state.composition).T
                heat_template_file = create_heat_template(st.session_state.unique_solutions, composition_df)
                st.download_button(
                    label="📥 Download Heat Template (Weighted Sum)",
                    data=heat_template_file,
                    file_name="weighted_sum_heat_template.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.warning("No weighted sum solutions available")
        
        with tab2:
            if st.session_state.e_constraint_solutions:
                st.markdown('<div class="pareto-container">', unsafe_allow_html=True)
                st.markdown('<div class="pareto-header">Adaptive ε-Constraint Method: Cost vs Emissions Trade-Off</div>', unsafe_allow_html=True)
                st.markdown('<div class="pareto-description">Solutions obtained by minimizing cost subject to emission constraints (ε) with adaptive refinement.</div>', unsafe_allow_html=True)
                
                plot_df = pd.DataFrame([{
                    "Solution": f"Solution {i+1}",
                    "Cost (INR/ton)": sol["Cost (INR/ton)"],
                    "Emissions (tCO₂)": sol["Emissions (tCO₂/tcs)"],
                    "Scrap %": sol["Scrap %"],
                    "Pre-Consumer %": sol["Pre-Consumer %"],
                    "Post-Consumer %": sol["Post-Consumer %"],
                    "Recycle Content %": sol["Recycle Content %"],
                    "Yield %": sol["Yield %"],
                    "Method": "Adaptive ε-Constraint"
                } for i, sol in enumerate(st.session_state.e_constraint_solutions)])
                
                fig = px.scatter(
                    plot_df,
                    x="Emissions (tCO₂)",
                    y="Cost (INR/ton)",
                    color="Recycle Content %",
                    symbol="Method",
                    hover_data=["Solution", "Scrap %", "Pre-Consumer %", "Post-Consumer %", "Yield %"],
                    labels={
                        "Emissions (tCO₂)": "Emissions (tCO₂/ton steel)",
                        "Cost (INR/ton)": "Production Cost (INR/ton steel)",
                        "Scrap %": "Scrap Percentage",
                        "Pre-Consumer %": "Pre-Consumer %",
                        "Post-Consumer %": "Post-Consumer %",
                        "Recycle Content %": "Recycle Content %",
                        "Yield %": "Yield Percentage",
                        "Method": "Optimization Method"
                    },
                    title="Adaptive ε-Constraint Method: Pareto Front",
                    template="plotly_white"
                )
                
                fig.update_traces(
                    marker=dict(size=10, symbol="diamond", line=dict(width=1, color='DarkSlateGrey')),
                    selector=dict(mode='markers')
                )
                
                fig.update_layout(
                    hoverlabel=dict(
                        bgcolor="white",
                        font_size=12,
                        font_family="Arial"
                    )
                )
                
                st.plotly_chart(fig, use_container_width=True)
                
                # Trade-off analysis for ε-constraint method
                st.markdown('<div class="tradeoff-analysis">', unsafe_allow_html=True)
                st.markdown('<div class="tradeoff-section-title">📊 Trade-Off Analysis</div>', unsafe_allow_html=True)
                
                if len(st.session_state.e_constraint_solutions) >= 2:
                    # Identify boundary solutions
                    ec_min_cost = min(st.session_state.e_constraint_solutions, key=lambda x: x["Cost (INR/ton)"])
                    ec_min_emission = min(st.session_state.e_constraint_solutions, key=lambda x: x["Emissions (tCO₂/tcs)"])
                    
                    # Calculate differences between boundary solutions
                    ec_cost_diff = ec_min_emission["Cost (INR/ton)"] - ec_min_cost["Cost (INR/ton)"]
                    ec_emission_diff = ec_min_cost["Emissions (tCO₂/tcs)"] - ec_min_emission["Emissions (tCO₂/tcs)"]
                    
                    st.markdown(f"""
                        <div class="tradeoff-section">
                            <p>The ε-constraint method reveals the following key solutions:</p>
                            <ul>
                                <li>The <strong>cheapest ε-constraint solution</strong> costs ₹{ec_min_cost["Cost (INR/ton)"]:,.0f}/ton with emissions of {ec_min_cost["Emissions (tCO₂/tcs)"]:.4f} tCO₂/ton</li>
                                <li>The <strong>lowest emission ε-constraint solution</strong> emits {ec_min_emission["Emissions (tCO₂/tcs)"]:.4f} tCO₂/ton at a cost of ₹{ec_min_emission["Cost (INR/ton)"]:,.0f}/ton</li>
                                <li>To reduce emissions by {ec_emission_diff:.4f} tCO₂/ton, costs increase by ₹{ec_cost_diff:,.0f}/ton ({ec_cost_diff/ec_min_cost["Cost (INR/ton)"]*100:.1f}%)</li>
                            </ul>
                        </div>
                    """, unsafe_allow_html=True)
                    
                    # Identify knee point for ε-constraint method
                    solutions_sorted = sorted(st.session_state.e_constraint_solutions, key=lambda x: x["Emissions (tCO₂/tcs)"])
                    slopes = []
                    for i in range(1, len(solutions_sorted)):
                        cost_diff = solutions_sorted[i]["Cost (INR/ton)"] - solutions_sorted[i-1]["Cost (INR/ton)"]
                        emission_diff = solutions_sorted[i]["Emissions (tCO₂/tcs)"] - solutions_sorted[i-1]["Emissions (tCO₂/tcs)"]
                        if emission_diff != 0:
                            slope = cost_diff / emission_diff
                            slopes.append((i, slope))
                    
                    if len(slopes) > 1:
                        max_slope_change_idx = 0
                        max_slope_change = 0
                        for i in range(1, len(slopes)):
                            slope_change = abs(slopes[i][1] - slopes[i-1][1])
                            if slope_change > max_slope_change:
                                max_slope_change = slope_change
                                max_slope_change_idx = i
                        
                        knee_point = solutions_sorted[slopes[max_slope_change_idx][0]]
                        prev_point = solutions_sorted[slopes[max_slope_change_idx][0]-1]
                        
                        # Compare material mixes to identify key differences
                        knee_materials = set(knee_point["mix"].keys())
                        prev_materials = set(prev_point["mix"].keys())
                        
                        added_materials = knee_materials - prev_materials
                        removed_materials = prev_materials - knee_materials
                        
                        st.markdown(f"""
                            <div class="tradeoff-section">
                                <div class="tradeoff-subsection-title">Knee Point Analysis</div>
                                <p>The knee point represents the optimal trade-off between cost and emissions:</p>
                                <ul>
                                    <li><strong>Cost:</strong> ₹{knee_point["Cost (INR/ton)"]:,.0f}/ton</li>
                                    <li><strong>Emissions:</strong> {knee_point["Emissions (tCO₂/tcs)"]:.4f} tCO₂/ton</li>
                                    <li><strong>Scrap Content:</strong> {knee_point["Scrap %"]:.1f}%</li>
                                </ul>
                                <p>At this point, the cost increases by ₹{(knee_point["Cost (INR/ton)"] - prev_point["Cost (INR/ton)"]):,.0f}/ton 
                                for an emission reduction of {(prev_point["Emissions (tCO₂/tcs)"] - knee_point["Emissions (tCO₂/tcs)"]):.4f} tCO₂/ton.</p>
                            </div>
                        """, unsafe_allow_html=True)
                        
                        if added_materials or removed_materials:
                            st.markdown('<div class="tradeoff-section">', unsafe_allow_html=True)
                            st.markdown('<div class="tradeoff-subsection-title">🔧 Key Material Changes at Knee Point</div>', unsafe_allow_html=True)
                            if added_materials:
                                st.markdown(f'<p>Materials added: {", ".join(added_materials)}</p>', unsafe_allow_html=True)
                                for mat in added_materials:
                                    mat_cost = st.session_state.cost.get(mat, 0)
                                    mat_emission = st.session_state.scope1.get(mat, 0) + st.session_state.scope3.get(mat, 0)
                                    st.markdown(f'<p>- {mat}: Cost ₹{mat_cost:,.0f}/ton, Emissions {mat_emission:.2f} tCO₂/ton</p>', unsafe_allow_html=True)
                            if removed_materials:
                                st.markdown(f'<p>Materials removed: {", ".join(removed_materials)}</p>', unsafe_allow_html=True)
                                for mat in removed_materials:
                                    mat_cost = st.session_state.cost.get(mat, 0)
                                    mat_emission = st.session_state.scope1.get(mat, 0) + st.session_state.scope3.get(mat, 0)
                                    st.markdown(f'<p>- {mat}: Cost ₹{mat_cost:,.0f}/ton, Emissions {mat_emission:.2f} tCO₂/ton</p>', unsafe_allow_html=True)
                            st.markdown('<p>This suggests that higher-cost, lower-emission materials are being substituted at this point.</p>', unsafe_allow_html=True)
                            st.markdown('</div>', unsafe_allow_html=True)
                
                st.markdown('</div>', unsafe_allow_html=True)  # Close tradeoff-analysis
                st.markdown('</div>', unsafe_allow_html=True)  # Close pareto-container
                
                # Create Excel download with detailed charge mixes
                excel_file = create_excel_download(st.session_state.e_constraint_solutions)
                st.download_button(
                    label="📥 Download ε-Constraint Solutions (Excel)",
                    data=excel_file,
                    file_name="e_constraint_solutions.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                # Create heat template download
                composition_df = pd.DataFrame(st.session_state.composition).T
                heat_template_file = create_heat_template(st.session_state.e_constraint_solutions, composition_df)
                st.download_button(
                    label="📥 Download Heat Template (ε-Constraint)",
                    data=heat_template_file,
                    file_name="e_constraint_heat_template.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.warning("No ε-constraint solutions available")
        
        st.subheader("📈 Combined View of Both Methods")
        
        if st.session_state.unique_solutions or st.session_state.e_constraint_solutions:
            combined_plot_df = pd.DataFrame([{
                "Solution": f"Solution {i+1}",
                "Cost (INR/ton)": sol["Cost (INR/ton)"],
                "Emissions (tCO₂)": sol["Emissions (tCO₂/tcs)"],
                "Scrap %": sol["Scrap %"],
                "Pre-Consumer %": sol["Pre-Consumer %"],
                "Post-Consumer %": sol["Post-Consumer %"],
                "Recycle Content %": sol["Recycle Content %"],
                "Yield %": sol["Yield %"],
                "Method": sol["Method"]
            } for i, sol in enumerate(all_solutions_sorted)])
            
            fig = px.scatter(
                combined_plot_df,
                x="Emissions (tCO₂)",
                y="Cost (INR/ton)",
                color="Method",
                symbol="Method",
                hover_data=["Solution", "Scrap %", "Pre-Consumer %", "Post-Consumer %", "Yield %"],
                labels={
                    "Emissions (tCO₂)": "Emissions (tCO₂/ton steel)",
                    "Cost (INR/ton)": "Production Cost (INR/ton steel)",
                    "Scrap %": "Scrap Percentage",
                    "Pre-Consumer %": "Pre-Consumer (Revert) %",
                    "Post-Consumer %": "Post-Consumer %",
                    "Recycle Content %": "Recycle Content %",
                    "Yield %": "Yield Percentage",
                    "Method": "Optimization Method"
                },
                title="Combined Pareto Front: Weighted Sum vs ε-Constraint Methods",
                template="plotly_white"
            )
            
            fig.update_traces(
                marker=dict(size=10, line=dict(width=1, color='DarkSlateGrey')),
                selector=dict(mode='markers')
            )
            
            fig.update_layout(
                hoverlabel=dict(
                    bgcolor="white",
                    font_size=12,
                    font_family="Arial"
                )
            )
            
            st.plotly_chart(fig, use_container_width=True)
            
            # Create Excel download with all solutions and detailed charge mixes
            excel_file = create_excel_download(all_solutions_sorted)
            st.download_button(
                label="📥 Download All Solutions (Excel)",
                data=excel_file,
                file_name="all_optimization_solutions.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            # Create heat template download for all solutions
            composition_df = pd.DataFrame(st.session_state.composition).T
            heat_template_file = create_heat_template(all_solutions_sorted, composition_df)
            st.download_button(
                label="📥 Download Heat Template (All Solutions)",
                data=heat_template_file,
                file_name="all_solutions_heat_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        st.subheader("🔍 Solution Explorer")
        
        tab1, tab2, tab3 = st.tabs(["📋 Summary Table", "📝 Detailed View", "🏭 Route Analysis"])
        
        with tab1:
            st.dataframe(
                pd.DataFrame([{
                    "Solution": f"Solution {i+1}",
                    "Method": sol["Method"],
                    "Cost (₹/ton)": sol["Cost (INR/ton)"],
                    "Emissions (tCO₂)": sol["Emissions (tCO₂/tcs)"],
                    "Scrap %": sol["Scrap %"],
                    "Pre-Consumer %": sol["Pre-Consumer %"],
                    "Post-Consumer %": sol["Post-Consumer %"],
                    "Recycle Content %": sol["Recycle Content %"],
                    "Yield %": sol["Yield %"],
                    "w1 (Cost Weight)": sol.get("w1", None),
                    "w2 (Emission Weight)": sol.get("w2", None)
                } for i, sol in enumerate(all_solutions_sorted)]).style.format({
                    "Cost (₹/ton)": "{:,.0f}",
                    "Emissions (tCO₂)": "{:.4f}",
                    "Scrap %": "{:.1f}%",
                    "Pre-Consumer %": "{:.1f}%",
                    "Post-Consumer %": "{:.1f}%",
                    "Recycle Content %": "{:.1f}%",
                    "Yield %": "{:.1f}%",
                    "w1 (Cost Weight)": "{:.2f}" if "w1" in all_solutions_sorted[0] else "",
                    "w2 (Emission Weight)": "{:.2f}" if "w2" in all_solutions_sorted[0] else ""
                }),
                use_container_width=True,
                height=400
            )
            
        with tab2:
            if all_solutions_sorted:
                selected_index = st.selectbox(
                    "Select a solution to view details:",
                    options=range(len(all_solutions_sorted)),
                    format_func=lambda x: f"Solution {x+1} (Cost: ₹{all_solutions_sorted[x]['Cost (INR/ton)']:,.0f}, Emissions: {all_solutions_sorted[x]['Emissions (tCO₂/tcs)']:.4f}, Method: {all_solutions_sorted[x]['Method']})"
                )
                
                selected_sol = all_solutions_sorted[selected_index]
                
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("### 📊 Solution Metrics")
                    st.markdown(f"""
                        <div class="metric-card">
                            <strong>Method:</strong> {selected_sol['Method']}<br>
                            <strong>Cost:</strong> ₹{selected_sol['Cost (INR/ton)']:,.2f} per ton LM
                        </div>
                        <div class="metric-card">
                            <strong>Emissions:</strong> {selected_sol['Emissions (tCO₂/tcs)']:.4f} tCO₂ per ton LM
                        </div>
                        <div class="metric-card">
                            <strong>Scrap Percentage:</strong> {selected_sol['Scrap %']:.2f}%
                        </div>
                    """, unsafe_allow_html=True)
                
                with col2:
                    st.markdown("### ♻️ Recycled Content")
                    st.markdown(f"""
                        <div class="metric-card">
                            <strong>Pre-Consumer %:</strong> {selected_sol['Pre-Consumer %']:.2f}%
                        </div>
                        <div class="metric-card">
                            <strong>Post-Consumer %:</strong> {selected_sol['Post-Consumer %']:.2f}%
                        </div>
                        <div class="metric-card">
                            <strong>Total Recycled Content(Pre + Post) %:</strong> {selected_sol['Recycle Content %']:.2f}%
                        </div>
                        {f'<div class="metric-card"><strong>Weights:</strong> Cost (w1) = {selected_sol["w1"]:.2f}, Emissions (w2) = {selected_sol["w2"]:.2f}</div>' if "w1" in selected_sol else ""}
                    """, unsafe_allow_html=True)
                
                st.markdown("### 🧾 Material Composition (kg per ton LM)")
                mix_details = selected_sol["mix"]
                if mix_details:
                    # Add total input row
                    total_input_kg = selected_sol["Total Input (kg)"]
                    total_row = pd.DataFrame({
                        "Material": ["TOTAL INPUT"],
                        "Quantity (kg)": [total_input_kg]
                    })
                    
                    mix_df = pd.DataFrame({
                        "Material": mix_details.keys(),
                        "Quantity (kg)": [qty for qty in mix_details.values()]
                    }).sort_values("Quantity (kg)", ascending=False)
                    
                    # Concatenate with total row
                    display_df = pd.concat([mix_df, total_row])
                    
                    st.dataframe(
                        display_df.style.format({"Quantity (kg)": "{:,.1f}"}),
                        use_container_width=True,
                        height=400
                    )
                    
                    # Add yield information
                    st.markdown(f"""
                        <div class="info-box">
                            <strong>Yield Information:</strong><br>
                            - Total Input Weight: {total_input_kg:,.1f} kg<br>
                            - Liquid Metal Output: 1,000.0 kg<br>
                            - Yield: {selected_sol['Yield %']:.1f}%
                        </div>
                    """, unsafe_allow_html=True)
                    
                    # Component-wise emissions breakdown
                    st.markdown("### 🌍 Component-wise Emissions Breakdown")
                    
                    # Create the emissions pie chart
                    emissions_pie_chart = create_component_emissions_pie_chart(
                        selected_sol["mix"], 
                        st.session_state.scope1, 
                        st.session_state.scope3
                    )
                    
                    st.plotly_chart(emissions_pie_chart, use_container_width=True)
                    
                    # Create a detailed emissions breakdown table
                    st.markdown("#### 📊 Detailed Emissions Contribution")
                    
                    emissions_data = []
                    for material, quantity_kg in selected_sol["mix"].items():
                        quantity_tons = quantity_kg / 1000
                        material_scope1 = st.session_state.scope1.get(material, 0)
                        material_scope3 = st.session_state.scope3.get(material, 0)
                        material_total_emission = material_scope1 + material_scope3
                        material_emission_contribution = material_total_emission * quantity_tons
                        
                        if material_emission_contribution > 0:
                            emissions_data.append({
                                "Material": material,
                                "Quantity (kg)": quantity_kg,
                                "Scope 1 Emission Factor": material_scope1,
                                "Scope 3 Emission Factor": material_scope3,
                                "Total Emission Factor": material_total_emission,
                                "Emission Contribution (tCO₂)": material_emission_contribution,
                                "Percentage of Total": (material_emission_contribution / selected_sol["Emissions (tCO₂/tcs)"]) * 100
                            })
                    
                    # Sort by emission contribution
                    emissions_df = pd.DataFrame(emissions_data).sort_values("Emission Contribution (tCO₂)", ascending=False)
                    
                    if not emissions_df.empty:
                        st.dataframe(
                            emissions_df.style.format({
                                "Quantity (kg)": "{:,.1f}",
                                "Scope 1 Emission Factor": "{:.4f}",
                                "Scope 3 Emission Factor": "{:.4f}",
                                "Total Emission Factor": "{:.4f}",
                                "Emission Contribution (tCO₂)": "{:.4f}",
                                "Percentage of Total": "{:.2f}%"
                            }),
                            use_container_width=True,
                            height=400
                        )
                    
                    # Add summary statistics
                    total_emission = selected_sol["Emissions (tCO₂/tcs)"]
                    top_5_contributors = emissions_df.head(5)
                    top_5_percentage = top_5_contributors["Percentage of Total"].sum()
                    
                    st.markdown(f"""
                        <div class="info-box">
                            <strong>Emissions Summary:</strong><br>
                            - Total Emissions: {total_emission:.4f} tCO₂/ton LM<br>
                            - Top 5 Materials Contribution: {top_5_percentage:.1f}% of total emissions<br>
                            - Number of Materials Contributing: {len(emissions_df)} materials
                        </div>
                    """, unsafe_allow_html=True)
                else:
                    st.warning("No materials in this solution.")
            else:
                st.warning("No solutions available to display")

        with tab3:
                 st.header(f"{st.session_state.selected_route} Route Analysis")
                 
                 # Initialize flux/fuel data if not already in session state
                 if st.session_state.flux_fuel_data is None:
                     st.session_state.flux_fuel_data = DEFAULT_FLUX_FUEL_DATA
                 
                 # Initialize cost data with default values (INR/kg for materials, INR/kWh for electricity)
                 if 'flux_fuel_cost' not in st.session_state:
                     st.session_state.flux_fuel_cost = {
                         "Lime": 21.0,
                         "Dolomite": 10.0,
                         "Electrode": 50.0,
                         "Coke": 23.0,
                         "Fluorspar": 46.0,
                         "LDO": 65.9,
                         "LSHS": 54.9,
                         "Propane": 66.1
                     }
                 
                 if 'electricity_cost' not in st.session_state:
                     st.session_state.electricity_cost = {
                         'Grid': 8.22,
                         'Renewable': 6.69
                     }
             
                 # Flux/Fuel Data Editor - Show default values at start
                 with st.expander("⚙️ Configure Process Parameters", expanded=True):
                     st.markdown("""
                         <div style="background-color:#f0f2f6;padding:15px;border-radius:5px;margin-bottom:20px;">
                             <h4 style="color:#2c3e50;margin-bottom:10px;">Process Configuration</h4>
                             <p style="color:#555;">Configure consumption values, emission factors, and costs for production parameters.</p>
                         </div>
                     """, unsafe_allow_html=True)
                     
                     # Show default values at the top
                     st.markdown("""
                         <div class="info-box">
                             <strong>Default Values:</strong><br>
                             • Flux Formers: Lime, Dolomite, Electrode, Coke, Fluorspar<br>
                             • Fuels: LDO, LSHS, Propane<br>
                             • Electricity: Grid (8.22 INR/kWh), Renewable (6.69 INR/kWh)<br>
                             • Emission Factors: As per industry standards
                         </div>
                     """, unsafe_allow_html=True)
                     
                     # Edit/Lock buttons
                     col1, col2 = st.columns([3, 1])
                     with col2:
                         if st.session_state.editing_flux_fuel:
                             if st.button("🔒 Save Configuration", key="lock_flux_button", type="primary"):
                                 st.session_state.editing_flux_fuel = False
                                 st.rerun()
                             if st.button("🔄 Reset Defaults", key="reset_flux_button"):
                                 st.session_state.flux_fuel_data = [item for item in DEFAULT_FLUX_FUEL_DATA 
                                                                   if item["Route"] == st.session_state.selected_route]
                                 st.rerun()
                         else:
                             if st.button("✏️ Edit Configuration", key="edit_flux_button", type="primary"):
                                 st.session_state.editing_flux_fuel = True
                                 st.rerun()
                     
                     if st.session_state.editing_flux_fuel:
                         # Flux Formers Section
                         with st.container():
                             st.markdown("""
                                 <div style="background-color:#e8f4f8;padding:10px 15px;border-radius:5px;margin-bottom:15px;">
                                     <h5 style="color:#2c3e50;margin:0;">Flux Formers</h5>
                                 </div>
                             """, unsafe_allow_html=True)
                             
                             flux_materials = ["Lime", "Dolomite", "Electrode", "Coke", "Fluorspar"]
                             flux_data = [item for item in st.session_state.flux_fuel_data 
                                         if item["Material"] in flux_materials and 
                                         item["Route"] == st.session_state.selected_route]
                             
                             # Header row
                             cols = st.columns([2, 2, 2, 2, 2, 1])
                             with cols[0]: st.markdown("**Material**")
                             with cols[1]: st.markdown("**Consumption**")
                             with cols[2]: st.markdown("**Scope 1 EF**")
                             with cols[3]: st.markdown("**Scope 3 EF**")
                             with cols[4]: st.markdown("**Cost**")
                             
                             for i, item in enumerate(flux_data):
                                 cols = st.columns([2, 2, 2, 2, 2, 1])
                                 with cols[0]:
                                     st.text_input("Material", value=item["Material"], 
                                                 key=f"flux_mat_{i}", label_visibility="collapsed")
                                 with cols[1]:
                                     st.number_input(
                                         "Consumption", 
                                         min_value=0.0, 
                                         max_value=0.5,
                                         value=float(item["Consumption"]), 
                                         step=0.0001,
                                         format="%.4f",
                                         key=f"flux_cons_{i}",
                                         label_visibility="collapsed"
                                     )
                                 with cols[2]:
                                     st.number_input(
                                         "Scope 1", 
                                         min_value=0.0, 
                                         max_value=10.0,
                                         value=float(item["Scope 1"]), 
                                         step=0.01,
                                         key=f"flux_s1_{i}",
                                         label_visibility="collapsed"
                                     )
                                 with cols[3]:
                                     st.number_input(
                                         "Scope 3", 
                                         min_value=0.0, 
                                         max_value=10.0,
                                         value=float(item["Scope 3"]), 
                                         step=0.01,
                                         key=f"flux_s3_{i}",
                                         label_visibility="collapsed"
                                     )
                                 with cols[4]:
                                     cost = st.number_input(
                                         "Cost", 
                                         min_value=0.0, 
                                         max_value=100.0,
                                         value=float(st.session_state.flux_fuel_cost.get(item["Material"], 0)), 
                                         step=0.1,
                                         key=f"flux_cost_{i}",
                                         label_visibility="collapsed"
                                     )
                                     st.session_state.flux_fuel_cost[item["Material"]] = cost
                                 with cols[5]:
                                     if st.button("🗑️", key=f"del_flux_{i}"):
                                         continue
                             
                             if st.button("➕ Add Flux Material", key="add_flux_material"):
                                 st.session_state.flux_fuel_data.append({
                                     "Material": "New Flux",
                                     "Route": st.session_state.selected_route,
                                     "Consumption": 0.0001,
                                     "Scope 1": 0.0,
                                     "Scope 3": 0.0
                                 })
                                 st.rerun()
                         
                         # Fuels Section
                         with st.container():
                             st.markdown("""
                                 <div style="background-color:#e8f4f8;padding:10px 15px;border-radius:5px;margin-bottom:15px;">
                                     <h5 style="color:#2c3e50;margin:0;">Fuels</h5>
                                 </div>
                             """, unsafe_allow_html=True)
                             
                             fuel_materials = ["LDO", "LSHS", "Propane"]
                             fuel_data = [item for item in st.session_state.flux_fuel_data 
                                         if item["Material"] in fuel_materials and 
                                         item["Route"] == st.session_state.selected_route]
                             
                             # Header row
                             cols = st.columns([2, 2, 2, 2, 2, 1])
                             with cols[0]: st.markdown("**Material**")
                             with cols[1]: st.markdown("**Consumption**")
                             with cols[2]: st.markdown("**Scope 1 EF**")
                             with cols[3]: st.markdown("**Scope 3 EF**")
                             with cols[4]: st.markdown("**Cost**")
                             
                             for i, item in enumerate(fuel_data):
                                 cols = st.columns([2, 2, 2, 2, 2, 1])
                                 with cols[0]:
                                     st.text_input("Material", value=item["Material"], 
                                                 key=f"fuel_mat_{i}", label_visibility="collapsed")
                                 with cols[1]:
                                     st.number_input(
                                         "Consumption", 
                                         min_value=0.0, 
                                         max_value=0.1,
                                         value=float(item["Consumption"]), 
                                         step=0.0001,
                                         format="%.4f",
                                         key=f"fuel_cons_{i}",
                                         label_visibility="collapsed"
                                     )
                                 with cols[2]:
                                     st.number_input(
                                         "Scope 1", 
                                         min_value=0.0, 
                                         max_value=10.0,
                                         value=float(item["Scope 1"]), 
                                         step=0.01,
                                         key=f"fuel_s1_{i}",
                                         label_visibility="collapsed"
                                     )
                                 with cols[3]:
                                     st.number_input(
                                         "Scope 3", 
                                         min_value=0.0, 
                                         max_value=10.0,
                                         value=float(item["Scope 3"]), 
                                         step=0.01,
                                         key=f"fuel_s3_{i}",
                                         label_visibility="collapsed"
                                     )
                                 with cols[4]:
                                     cost = st.number_input(
                                         "Cost", 
                                         min_value=0.0, 
                                         max_value=100.0,
                                         value=float(st.session_state.flux_fuel_cost.get(item["Material"], 0)), 
                                         step=0.1,
                                         key=f"fuel_cost_{i}",
                                         label_visibility="collapsed"
                                     )
                                     st.session_state.flux_fuel_cost[item["Material"]] = cost
                                 with cols[5]:
                                     if st.button("🗑️", key=f"del_fuel_{i}"):
                                         continue
                             
                             if st.button("➕ Add Fuel Material", key="add_fuel_material"):
                                 st.session_state.flux_fuel_data.append({
                                     "Material": "New Fuel",
                                     "Route": st.session_state.selected_route,
                                     "Consumption": 0.0001,
                                     "Scope 1": 0.0,
                                     "Scope 3": 0.0
                                 })
                                 st.rerun()
                         
                         # Electricity Section
                         with st.container():
                             st.markdown("""
                                 <div style="background-color:#e8f4f8;padding:10px 15px;border-radius:5px;margin-bottom:15px;">
                                     <h5 style="color:#2c3e50;margin:0;">Electricity Parameters</h5>
                                 </div>
                             """, unsafe_allow_html=True)
                             
                             cols = st.columns(2)
                             with cols[0]:
                                 st.markdown("**Emission Factors (kg CO₂/kWh)**")
                                 re_scope2 = st.number_input(
                                     "RE Scope 2", 
                                     min_value=0.0, 
                                     max_value=1.0,
                                     value=float(st.session_state.electricity_ef.get('RE_scope2', 0.00)), 
                                     step=0.001,
                                     format="%.3f",
                                     key="re_scope2"
                                 )
                                 re_scope3 = st.number_input(
                                     "RE Scope 3", 
                                     min_value=0.0, 
                                     max_value=1.0,
                                     value=float(st.session_state.electricity_ef.get('RE_scope3', 0.31)), 
                                     step=0.001,
                                     format="%.3f",
                                     key="re_scope3"
                                 )
                                 grid_scope2 = st.number_input(
                                     "Grid Scope 2", 
                                     min_value=0.0, 
                                     max_value=1.0,
                                     value=float(st.session_state.electricity_ef.get('Grid_scope2', 0.727)), 
                                     step=0.001,
                                     format="%.3f",
                                     key="grid_scope2"
                                 )
                                 grid_scope3 = st.number_input(
                                     "Grid Scope 3", 
                                     min_value=0.0, 
                                     max_value=1.0,
                                     value=float(st.session_state.electricity_ef.get('Grid_scope3', 0.31)), 
                                     step=0.001,
                                     format="%.3f",
                                     key="grid_scope3"
                                 )
                             
                             with cols[1]:
                                 st.markdown("**Costs (INR/kWh)**")
                                 re_cost = st.number_input(
                                     "Renewable Cost", 
                                     min_value=0.0, 
                                     max_value=20.0,
                                     value=float(st.session_state.electricity_cost.get('Renewable', 6.69)), 
                                     step=0.01,
                                     format="%.2f",
                                     key="re_cost"
                                 )
                                 grid_cost = st.number_input(
                                     "Grid Cost", 
                                     min_value=0.0, 
                                     max_value=20.0,
                                     value=float(st.session_state.electricity_cost.get('Grid', 8.22)), 
                                     step=0.01,
                                     format="%.2f",
                                     key="grid_cost"
                                 )
                             
                             # Update session state
                             st.session_state.electricity_ef = {
                                 'RE_scope2': re_scope2,
                                 'RE_scope3': re_scope3,
                                 'Grid_scope2': grid_scope2,
                                 'Grid_scope3': grid_scope3
                             }
                             st.session_state.electricity_cost = {
                                 'Renewable': re_cost,
                                 'Grid': grid_cost
                             }
                     
                     else:
                         # Display non-editable tables
                         # Flux Formers Section
                         with st.container():
                             st.markdown("""
                                 <div style="background-color:#e8f4f8;padding:10px 15px;border-radius:5px;margin-bottom:15px;">
                                     <h5 style="color:#2c3e50;margin:0;">Flux Formers</h5>
                                 </div>
                             """, unsafe_allow_html=True)
                             
                             flux_materials = ["Lime", "Dolomite", "Electrode", "Coke", "Fluorspar"]
                             flux_df = pd.DataFrame([item for item in st.session_state.flux_fuel_data 
                                                   if item["Material"] in flux_materials and 
                                                   item["Route"] == st.session_state.selected_route])
                             
                             if not flux_df.empty:
                                 # Add cost column
                                 flux_df["Cost (INR/kg)"] = flux_df["Material"].map(st.session_state.flux_fuel_cost)
                                 
                                 # Format consumption
                                 flux_df["Consumption"] = flux_df["Consumption"].apply(lambda x: f"{x:.4f}")
                                 
                                 st.dataframe(
                                     flux_df[["Material", "Consumption", "Scope 1", "Scope 3", "Cost (INR/kg)"]]
                                     .rename(columns={
                                         "Consumption": "Consumption (ton/ton)",
                                         "Scope 1": "Scope 1 (kg CO₂/kg)",
                                         "Scope 3": "Scope 3 (kg CO₂/kg)"
                                     }),
                                     use_container_width=True,
                                     hide_index=True
                                 )
                             else:
                                 st.warning("No flux formers configured for this route")
                         
                         # Fuels Section
                         with st.container():
                             st.markdown("""
                                 <div style="background-color:#e8f4f8;padding:10px 15px;border-radius:5px;margin-bottom:15px;">
                                     <h5 style="color:#2c3e50;margin:0;">Fuels</h5>
                                 </div>
                             """, unsafe_allow_html=True)
                             
                             fuel_materials = ["LDO", "LSHS", "Propane"]
                             fuel_df = pd.DataFrame([item for item in st.session_state.flux_fuel_data 
                                                   if item["Material"] in fuel_materials and 
                                                   item["Route"] == st.session_state.selected_route])
                             
                             if not fuel_df.empty:
                                 # Add cost column
                                 fuel_df["Cost (INR/kg)"] = fuel_df["Material"].map(st.session_state.flux_fuel_cost)
                                 
                                 # Format consumption
                                 fuel_df["Consumption"] = fuel_df["Consumption"].apply(lambda x: f"{x:.4f}")
                                 
                                 st.dataframe(
                                     fuel_df[["Material", "Consumption", "Scope 1", "Scope 3", "Cost (INR/kg)"]]
                                     .rename(columns={
                                         "Consumption": "Consumption (ton/ton)",
                                         "Scope 1": "Scope 1 (kg CO₂/kg)",
                                         "Scope 3": "Scope 3 (kg CO₂/kg)"
                                     }),
                                     use_container_width=True,
                                     hide_index=True
                                 )
                             else:
                                 st.warning("No fuels configured for this route")
                         
                         # Electricity Section
                         with st.container():
                             st.markdown("""
                                 <div style="background-color:#e8f4f8;padding:10px 15px;border-radius:5px;margin-bottom:15px;">
                                     <h5 style="color:#2c3e50;margin:0;">Electricity Parameters</h5>
                                 </div>
                             """, unsafe_allow_html=True)
                             
                             # Combined Table for Electricity Parameters
                             elec_params = pd.DataFrame({
                                 "Type": ["Renewable Energy", "Grid Electricity"],
                                 "Scope 2 (kg CO₂/kWh)": [
                                     st.session_state.electricity_ef['RE_scope2'],
                                     st.session_state.electricity_ef['Grid_scope2']
                                 ],
                                 "Scope 3 (kg CO₂/kWh)": [
                                     st.session_state.electricity_ef['RE_scope3'],
                                     st.session_state.electricity_ef['Grid_scope3']
                                 ],
                                 "Total Emissions (kg CO₂/kWh)": [
                                     st.session_state.electricity_ef['RE_scope2'] + st.session_state.electricity_ef['RE_scope3'],
                                     st.session_state.electricity_ef['Grid_scope2'] + st.session_state.electricity_ef['Grid_scope3']
                                 ],
                                 "Cost (INR/kWh)": [
                                     st.session_state.electricity_cost['Renewable'],
                                     st.session_state.electricity_cost['Grid']
                                 ]
                             })
                             
                             st.dataframe(
                                 elec_params.style.format({
                                     "Scope 2 (kg CO₂/kWh)": "{:.3f}",
                                     "Scope 3 (kg CO₂/kWh)": "{:.3f}",
                                     "Total Emissions (kg CO₂/kWh)": "{:.3f}",
                                     "Cost (INR/kWh)": "{:.2f}"
                                 }),
                                 use_container_width=True,
                                 hide_index=True
                             )
             
                 # Route Parameters Section
                 st.markdown("""
                     <div style="background-color:#f0f2f6;padding:15px;border-radius:5px;margin-bottom:20px;">
                         <h4 style="color:#2c3e50;margin-bottom:10px;">Production Parameters</h4>
                     </div>
                 """, unsafe_allow_html=True)
                 
                 col1, col2 = st.columns(2)
                 with col1:
                     route_elec = st.number_input(
                         f"{st.session_state.selected_route} Electricity (kWh/ton product)", 
                         min_value=300.0, 
                         max_value=900.0, 
                         value=612.0 if st.session_state.selected_route == "EAF" else 710.0, 
                         key="route_elec"
                     )
                 with col2:
                     route_renewable = st.slider(
                         f"{st.session_state.selected_route} Renewable %", 
                         0, 100, 
                         30 if st.session_state.selected_route == "EAF" else 20, 
                         key="route_renewable"
                     )
                 
                 casting_yield = st.number_input(
                     f"{st.session_state.selected_route} Liquid to Casting Yield (%)",
                     min_value=50.0,
                     max_value=100.0,
                     value=96.0 if st.session_state.selected_route == "EAF" else 95.0,
                     step=0.1,
                     format="%.1f",
                     key="casting_yield"
                 ) / 100  # Convert to decimal
             
                 # Get emission factors from session state
                 EF_RE_scope2 = st.session_state.electricity_ef.get('RE_scope2', 0.00)
                 EF_RE_scope3 = st.session_state.electricity_ef.get('RE_scope3', 0.31)
                 EF_Grid_scope2 = st.session_state.electricity_ef.get('Grid_scope2', 0.727)
                 EF_Grid_scope3 = st.session_state.electricity_ef.get('Grid_scope3', 0.31)
                 
                 EF_RE_total = EF_RE_scope2 + EF_RE_scope3
                 EF_Grid_total = EF_Grid_scope2 + EF_Grid_scope3
                 
                 # Get cost factors from session state
                 Cost_RE = st.session_state.electricity_cost.get('Renewable', 6.69)
                 Cost_Grid = st.session_state.electricity_cost.get('Grid', 8.22)
             
                 # Function to calculate fixed emissions (Scope 1 + Scope 3) in ton CO₂/ton product
                 def compute_fixed_emissions(route):
                     total_kg_CO2 = 0
                     for item in st.session_state.flux_fuel_data:
                         if item["Route"] == route:
                             consumption = item.get("Consumption", 0)  # ton/ton product
                             scope1 = item.get("Scope 1", 0)  # kg CO₂/kg material
                             scope3 = item.get("Scope 3", 0)  # kg CO₂/kg material
                             total = consumption * (scope1 + scope3) * 1000  # Convert to kg CO₂/ton product
                             total_kg_CO2 += total
                     return total_kg_CO2 / 1000  # Convert from kg to ton CO₂/ton product
                 
                 # Function to calculate fixed costs in INR/ton product
                 def compute_fixed_costs(route):
                     total_cost = 0
                     for item in st.session_state.flux_fuel_data:
                         if item["Route"] == route:
                             mat = item.get("Material", "Unknown")
                             consumption = item.get("Consumption", 0)  # ton/ton product
                             cost = st.session_state.flux_fuel_cost.get(mat, 0)  # INR/kg
                             total_cost += consumption * cost * 1000  # Convert to INR/ton product
                     return total_cost
                 
                 # Calculate Route Emissions and Costs
                 def calculate_route_parameters(route, elec_consumption, renewable_pct):
                     flux_fuel_emis = compute_fixed_emissions(route)  # ton CO₂/ton product
                     flux_fuel_cost = compute_fixed_costs(route)  # INR/ton product
                     
                     grid_portion = (100 - renewable_pct)/100
                     renew_portion = renewable_pct/100
                     
                     # Calculate electricity emissions
                     elec_emis = elec_consumption * (
                         grid_portion * EF_Grid_total + 
                         renew_portion * EF_RE_total
                     ) / 1000  # Convert from kg to ton CO₂/ton product
                     
                     # Calculate electricity costs
                     elec_cost = elec_consumption * (
                         grid_portion * Cost_Grid + 
                         renew_portion * Cost_RE
                     )  # INR/ton product
                     
                     total_emis = flux_fuel_emis + elec_emis
                     total_cost = flux_fuel_cost + elec_cost
                     
                     return {
                         'Total_Emissions': total_emis,
                         'Total_Cost': total_cost,
                         'Electricity_Emissions': elec_emis,
                         'Electricity_Cost': elec_cost,
                         'Flux_Fuel_Emissions': flux_fuel_emis,
                         'Flux_Fuel_Cost': flux_fuel_cost,
                         'RE_Percent': renewable_pct
                     }
             
                 # Enhance Solutions with Production Emissions and Costs
                 def enhance_solutions(solutions, route_params, casting_yield):
                     enhanced = []
                     for i, sol in enumerate(solutions):
                         # Calculate scaling factors
                         scale = 1 / casting_yield if casting_yield > 0 else 1
                         
                         # Create solution with default values for missing keys
                         solution_data = {
                             "Cost (INR/ton)": sol.get("Cost (INR/ton)", 0),
                             "Emissions (tCO₂/tcs)": sol.get("Emissions (tCO₂/tcs)", 0),
                             "Scrap %": sol.get("Scrap %", 0),
                             "Pre-Consumer %": sol.get("Pre-Consumer %", 0),
                             "Post-Consumer %": sol.get("Post-Consumer %", 0),
                             "Recycle Content %": sol.get("Recycle Content %", 0),
                             "Yield %": sol.get("Yield %", 0),
                             "Total Input (kg)": sol.get("Total Input (kg)", 0),
                             "Method": sol.get("Method", "Unknown"),
                             "mix": sol.get("mix", {}),
                             "Solution": f"Solution {i+1}",
                             "Total_Emissions": (sol.get("Emissions (tCO₂/tcs)", 0) * scale) + route_params['Total_Emissions'],
                             "Total_Cost": (sol.get("Cost (INR/ton)", 0) * scale) + route_params['Total_Cost'],
                             "Production_Emissions": route_params['Total_Emissions'],
                             "Production_Cost": route_params['Total_Cost'],
                             "Electricity_Emissions": route_params['Electricity_Emissions'],
                             "Electricity_Cost": route_params['Electricity_Cost'],
                             "Flux_Fuel_Emissions": route_params['Flux_Fuel_Emissions'],
                             "Flux_Fuel_Cost": route_params['Flux_Fuel_Cost'],
                             "Route": st.session_state.selected_route,
                             "RE_Percent": route_params['RE_Percent'],
                             "Casting_Yield": casting_yield
                         }
                         
                         enhanced.append(solution_data)
                     return enhanced
             
                 current_route_params = calculate_route_parameters(st.session_state.selected_route, route_elec, route_renewable)
                 
                 enhanced_solutions = enhance_solutions(
                     all_solutions_sorted,
                     current_route_params,
                     casting_yield
                 )
                 
                 plot_df = pd.DataFrame(enhanced_solutions)
                 
                 # Cost vs Emissions Graph
                 st.markdown("""
                     <div style="background-color:#f0f2f6;padding:15px;border-radius:5px;margin-bottom:20px;">
                         <h4 style="color:#2c3e50;margin-bottom:10px;">Cost vs Emissions Analysis</h4>
                     </div>
                 """, unsafe_allow_html=True)
                 
                 fig = px.scatter(
                     plot_df,
                     x="Total_Emissions",
                     y="Total_Cost",
                     color="Method",
                     symbol="Method",
                     hover_data=["Solution", "Scrap %", "Casting_Yield", "RE_Percent"],
                     labels={
                         "Total_Emissions": "Total Emissions (ton CO₂/ton product)",
                         "Total_Cost": "Total Cost (INR/ton product)",
                         "Method": "Optimization Method"
                     },
                     template="plotly_white"
                 )
                 
                 fig.update_layout(
                     plot_bgcolor='white',
                     paper_bgcolor='white',
                     xaxis=dict(
                         title_font=dict(size=12),
                         tickfont=dict(size=10),
                         gridcolor='lightgray'
                     ),
                     yaxis=dict(
                         title_font=dict(size=12),
                         tickfont=dict(size=10),
                         gridcolor='lightgray'
                     ),
                     legend=dict(
                         title_font=dict(size=10),
                         font=dict(size=9)
                     )
                 )
                 
                 st.plotly_chart(fig, use_container_width=True)
             
                 # Create tabs for Cost and Emission Analysis
                 cost_tab, emis_tab = st.tabs(["💰 Cost Analysis", "🌍 Emission Analysis"])
                 
                 with cost_tab:
                     st.markdown("""
                         <div style="background-color:#f0f2f6;padding:15px;border-radius:5px;margin-bottom:20px;">
                             <h4 style="color:#2c3e50;margin-bottom:10px;">Cost Breakdown</h4>
                         </div>
                     """, unsafe_allow_html=True)
                     
                     # Cost breakdown for a selected solution
                     solution_options = [
                         f"Solution {i+1} (Cost: ₹{sol['Cost (INR/ton)']:,.0f}, Emissions: {sol['Emissions (tCO₂/tcs)']:.3f}, Method: {sol.get('Method', 'Unknown')})" 
                         for i, sol in enumerate(all_solutions_sorted)
                     ]
                     
                     selected_solution = st.selectbox(
                         "Select Solution for Cost Breakdown:",
                         options=solution_options,
                         index=0
                     )
                     selected_index = solution_options.index(selected_solution)
                     
                     selected_sol = all_solutions_sorted[selected_index]
                     scale = 1 / casting_yield if casting_yield > 0 else 1
                     
                     # Calculate cost components - SEPARATE FLUX AND FUEL
                     raw_mat_cost = selected_sol.get("Cost (INR/ton)", 0) * scale
                     
                     # Separate flux and fuel costs
                     flux_cost = 0
                     fuel_cost = 0
                     
                     for item in st.session_state.flux_fuel_data:
                         if item["Route"] == st.session_state.selected_route:
                             mat = item.get("Material", "Unknown")
                             consumption = item.get("Consumption", 0)  # ton/ton product
                             cost_per_kg = st.session_state.flux_fuel_cost.get(mat, 0)  # INR/kg
                             item_cost = consumption * cost_per_kg * 1000  # Convert to INR/ton product
                             
                             if mat in ["Lime", "Dolomite", "Electrode", "Coke", "Fluorspar"]:
                                 flux_cost += item_cost
                             elif mat in ["LDO", "LSHS", "Propane"]:
                                 fuel_cost += item_cost
                     
                     elec_cost = current_route_params['Electricity_Cost']
                     total_cost = raw_mat_cost + flux_cost + fuel_cost + elec_cost
                     
                     # Create cost breakdown data for pie chart
                     cost_data = {
                         'Raw Materials': raw_mat_cost,
                         'Flux': flux_cost,
                         'Fuel': fuel_cost,
                         'Electricity': elec_cost
                     }
                     
                     # Filter out zero values
                     cost_data = {k: v for k, v in cost_data.items() if v > 0}
                     
                     # Create pie chart
                     if cost_data:
                         fig_cost = px.pie(
                             values=list(cost_data.values()),
                             names=list(cost_data.keys()),
                             title="Cost Breakdown (INR/ton product)",
                             color=list(cost_data.keys()),
                             color_discrete_map={
                                 'Raw Materials': '#1f77b4',
                                 'Flux': '#ff7f0e',
                                 'Fuel': '#2ca02c',
                                 'Electricity': '#d62728'
                             }
                         )
                         
                         fig_cost.update_traces(
                             textposition='inside',
                             textinfo='percent+label',
                             hovertemplate="%{label}: ₹%{value:,.2f}<br>%{percent}<extra></extra>"
                         )
                         
                         fig_cost.update_layout(
                             showlegend=True,
                             legend=dict(
                                 orientation="h",
                                 yanchor="bottom",
                                 y=-0.2,
                                 xanchor="center",
                                 x=0.5
                             ),
                             height=400
                         )
                         
                         st.plotly_chart(fig_cost, use_container_width=True)
                         
                         # Display total cost
                         st.markdown(f"""
                             <div class="metric-card">
                                 <strong>Total Cost:</strong> ₹{total_cost:,.2f} per ton product
                             </div>
                         """, unsafe_allow_html=True)
                 
                 with emis_tab:
                     st.markdown("""
                         <div style="background-color:#f0f2f6;padding:15px;border-radius:5px;margin-bottom:20px;">
                             <h4 style="color:#2c3e50;margin-bottom:10px;">Emission Breakdown</h4>
                         </div>
                     """, unsafe_allow_html=True)
                     
                     # Emission breakdown for the selected solution
                     selected_sol = all_solutions_sorted[selected_index]
                     scale = 1 / casting_yield if casting_yield > 0 else 1
                     
                     # Calculate emission components - SEPARATE FLUX AND FUEL
                     raw_mat_emis = selected_sol.get("Emissions (tCO₂/tcs)", 0) * scale
                     
                     # Separate flux and fuel emissions
                     flux_emis = 0
                     fuel_emis = 0
                     
                     for item in st.session_state.flux_fuel_data:
                         if item["Route"] == st.session_state.selected_route:
                             consumption = item.get("Consumption", 0)  # ton/ton product
                             scope1 = item.get("Scope 1", 0)  # kg CO₂/kg material
                             scope3 = item.get("Scope 3", 0)  # kg CO₂/kg material
                             item_emis = consumption * (scope1 + scope3) * 1000 / 1000  # Convert to ton CO₂/ton product
                             
                             if item["Material"] in ["Lime", "Dolomite", "Electrode", "Coke", "Fluorspar"]:
                                 flux_emis += item_emis
                             elif item["Material"] in ["LDO", "LSHS", "Propane"]:
                                 fuel_emis += item_emis
                     
                     elec_emis = current_route_params['Electricity_Emissions']
                     total_emis = raw_mat_emis + flux_emis + fuel_emis + elec_emis
                     
                     # Create emission breakdown data for pie chart
                     emission_data = {
                         'Raw Materials': raw_mat_emis,
                         'Flux': flux_emis,
                         'Fuel': fuel_emis,
                         'Electricity': elec_emis
                     }
                     
                     # Filter out zero values
                     emission_data = {k: v for k, v in emission_data.items() if v > 0}
                     
                     # Create pie chart
                     if emission_data:
                         fig_emis = px.pie(
                             values=list(emission_data.values()),
                             names=list(emission_data.keys()),
                             title="Emission Breakdown (ton CO₂/ton product)",
                             color=list(emission_data.keys()),
                             color_discrete_map={
                                 'Raw Materials': '#1f77b4',
                                 'Flux': '#ff7f0e',
                                 'Fuel': '#2ca02c',
                                 'Electricity': '#d62728'
                             }
                         )
                         
                         fig_emis.update_traces(
                             textposition='inside',
                             textinfo='percent+label',
                             hovertemplate="%{label}: %{value:.4f} tCO₂<br>%{percent}<extra></extra>"
                         )
                         
                         fig_emis.update_layout(
                             showlegend=True,
                             legend=dict(
                                 orientation="h",
                                 yanchor="bottom",
                                 y=-0.2,
                                 xanchor="center",
                                 x=0.5
                             ),
                             height=400
                         )
                         
                         st.plotly_chart(fig_emis, use_container_width=True)
                         
                         # Display total emissions
                         st.markdown(f"""
                             <div class="metric-card">
                                 <strong>Total Emissions:</strong> {total_emis:.4f} ton CO₂ per ton product
                             </div>
                         """, unsafe_allow_html=True)
                 
                 # Sensitivity Analysis Section
                 st.markdown("""
                     <div style="background-color:#f0f2f6;padding:15px;border-radius:5px;margin-bottom:20px;">
                         <h4 style="color:#2c3e50;margin-bottom:10px;">Electricity Sensitivity Analysis</h4>
                     </div>
                 """, unsafe_allow_html=True)
                 
                 # Dropdown to select solution
                 solution_options = [
                     f"Solution {i+1} (Cost: ₹{sol['Cost (INR/ton)']:,.0f}, Emissions: {sol['Emissions (tCO₂/tcs)']:.3f}, Method: {sol.get('Method', 'Unknown')})" 
                     for i, sol in enumerate(all_solutions_sorted)
                 ]
                 
                 selected_solution = st.selectbox(
                     "Select Solution for Sensitivity Analysis:",
                     options=solution_options,
                     index=0
                 )
                 selected_index = solution_options.index(selected_solution)
                 
                 # Create a range of renewable percentages to analyze
                 renewable_percentages = list(range(0, 101, 10))
                 
                 # Calculate emissions and costs for selected solution at different renewable percentages
                 sensitivity_data = []
                 for pct in renewable_percentages:
                     # Calculate electricity parameters for this renewable percentage
                     grid_portion = (100 - pct)/100
                     renew_portion = pct/100
                     
                     # Emissions
                     elec_emis = route_elec * (
                         grid_portion * EF_Grid_total + 
                         renew_portion * EF_RE_total
                     ) / 1000  # Convert from kg to ton CO₂/ton product
                     
                     # Costs
                     elec_cost = route_elec * (
                         grid_portion * Cost_Grid + 
                         renew_portion * Cost_RE
                     )  # INR/ton product
                     
                     # Total values (with scaled raw material values)
                     scale = 1 / casting_yield if casting_yield > 0 else 1
                     sol = all_solutions_sorted[selected_index]
                     total_emis = (sol.get("Emissions (tCO₂/tcs)", 0) * scale) + current_route_params['Flux_Fuel_Emissions'] + elec_emis
                     total_cost = (sol.get("Cost (INR/ton)", 0) * scale) + current_route_params['Flux_Fuel_Cost'] + elec_cost
                     
                     sensitivity_data.append({
                         "Renewable %": pct,
                         "Total Emissions (tCO₂)": total_emis,
                         "Total Cost (INR)": total_cost,
                         "Electricity Cost (INR)": elec_cost,
                         "Electricity Emissions (tCO₂)": elec_emis
                     })
                 
                 sensitivity_df = pd.DataFrame(sensitivity_data)
                 
                 # Create interactive plot for cost
                 fig_cost = px.line(
                     sensitivity_df,
                     x="Renewable %",
                     y="Total Cost (INR)",
                     labels={
                         "Total Cost (INR)": "Total Cost (INR/ton product)"
                     },
                     template="plotly_white"
                 )
                 
                 fig_cost.update_layout(
                     plot_bgcolor='white',
                     paper_bgcolor='white',
                     xaxis=dict(
                         title="Renewable Energy Percentage (%)",
                         title_font=dict(size=12),
                         tickfont=dict(size=10),
                         gridcolor='lightgray'
                     ),
                     yaxis=dict(
                         title="Total Cost (INR/ton product)",
                         title_font=dict(size=12),
                         tickfont=dict(size=10),
                         gridcolor='lightgray'
                     )
                 )
                 
                 st.plotly_chart(fig_cost, use_container_width=True)
                 
                 # Create interactive plot for emissions
                 fig_emis = px.line(
                     sensitivity_df,
                     x="Renewable %",
                     y="Total Emissions (tCO₂)",
                     labels={
                         "Total Emissions (tCO₂)": "Total Emissions (ton CO₂/ton product)"
                     },
                     template="plotly_white"
                 )
                 
                 fig_emis.update_layout(
                     plot_bgcolor='white',
                     paper_bgcolor='white',
                     xaxis=dict(
                         title="Renewable Energy Percentage (%)",
                         title_font=dict(size=12),
                         tickfont=dict(size=10),
                         gridcolor='lightgray'
                     ),
                     yaxis=dict(
                         title="Total Emissions (ton CO₂/ton product)",
                         title_font=dict(size=12),
                         tickfont=dict(size=10),
                         gridcolor='lightgray'
                     )
                 )
                 
                 st.plotly_chart(fig_emis, use_container_width=True)
        
    # Navigation buttons
    col1, col2 = st.columns([1, 1])
    with col1:
        if st.button("⬅️ Back", key="back_to_optimize", type="secondary"):
            st.session_state.current_step = 4
            st.rerun()
    with col2:
        st.button("➡️ Next", disabled=True, help="This is the final step")

# Welcome screen when no data is loaded
else:
    # Stylish header
    st.markdown("""
        <div class="app-header">
            <h1 class="app-title">Stainless Steel Charge-Mix Optimizer</h1>
            <p class="app-subtitle">Find optimal material combinations balancing cost and emissions</p>
        </div>
    """, unsafe_allow_html=True)
    
    st.markdown("""
        <div class="info-box">
            <h3>👋 Welcome to the Steel Charge Mix Optimizer</h3>
            <p>To get started, please upload your Excel input file containing:</p>
            <ul>
                <li><strong>Materials_data</strong> sheet with material properties</li>
                <li><strong>Composition_constraint</strong> sheet with material compositions</li>
            </ul>
            <p>The app will help you find optimal charge mixes that balance cost and emissions while meeting all requirements.</p>
        </div>
    """, unsafe_allow_html=True)
    
    # Add example file download
    with st.expander("💡 Need an example file?"):
        st.markdown("""
            Download this example Excel file to see the required format:
            [Example Input File](https://example.com/steel_optimizer_template.xlsx)
        """)